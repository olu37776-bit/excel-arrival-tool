from __future__ import annotations

from decimal import Decimal
from pathlib import Path
from tempfile import TemporaryDirectory
import json
import os
import shutil
import subprocess
import unittest
from zipfile import ZipFile
from xml.etree import ElementTree as ET

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from revenue_tool.adapters.excel_reader import ExcelInputAdapter
from revenue_tool.adapters.excel_writer import ExcelOutputAdapter
from revenue_tool.adapters.final_revenue_formulas import final_formulas
from revenue_tool.config import load_config
from revenue_tool.domain.models import BaseRow, IssueLog
from revenue_tool.services.final_revenue import (
    calculate_final_values, FINAL_FIELD_SOURCES, YEAR_REQUIRED_HINT,
    INVALID_MONTH_HINT,
)
from tests.test_pipeline import CONFIG, _run, _write_sources, _base_rows
from tests.test_manual_revenue_forecast import (
    _set_manual_inputs, _remove_fields_from_result,
    _insert_noncritical_column_and_move_to_end, _rename_metadata_backed_field,
)


def formula_cases():
    """Same cases consumed by Python and an independent spreadsheet engine."""
    base = dict(revenue_segment="订未发", revenue_forecast=100,
                revenue_month_rpd="2026-09", revenue_month_cpd="2026-10")
    cases = [base]
    for value in (None, "", "   ", "(空白)", "VALUE", "#VALUE", "#VALUE!",
                  "特殊处理", "9月确认", 0, False, True, 123, "Y", "N"):
        cases.append(dict(base, manual_revenue_segment=value))
    for value in (None, 0, 20.125, -20.125, "123.456", "(123.45)", "(12x", "abc", False,
                  "9/10", "12:30", "$10", "10%", "1e2", "1,234.565"):
        cases.append(dict(base, manual_revenue_month=value, manual_adjust_flag="N"))
    for mode in ("rpd", "cpd"):
        field = f"manual_revenue_forecast_{mode}"
        for value in (None, "", "(空白)", "VALUE", "9", "09", 9, "9月",
                      "09月", "9月份", "10", "10月", "2026-9", "2026-09",
                      "2026-13", "13月", 0, False, "abc", "9-10", "1e1"):
            cases.append(dict(base, **{field: value}))
        for reference_month in range(1, 13):
            for manual_month in range(1, 13):
                cases.append(dict(base, **{
                    f"revenue_month_{mode}": f"2026-{reference_month:02d}",
                    field: f"{manual_month}月",
                }))
        other = "cpd" if mode == "rpd" else "rpd"
        cases.extend([
            dict(base, **{f"revenue_month_{mode}": None,
                          f"revenue_month_{other}": "2026-12", field: "1月"}),
            dict(base, revenue_month_rpd=None, revenue_month_cpd=None,
                 **{field: "9月"}),
            dict(base, revenue_month_rpd=None, revenue_month_cpd=None,
                 **{field: "2026-9"}),
        ])
    cases.append(dict(base, revenue_month_rpd=None, revenue_month_cpd=None,
                      revenue_segment=None))
    return cases


class FinalRevenueTest(unittest.TestCase):
    def test_priority_and_falsy_inputs_do_not_depend_on_flag(self):
        for flag in (None, "N", "Y"):
            for value in ("特殊处理", 0, False, 123, True):
                values = dict(revenue_segment="订未发", manual_revenue_segment=value,
                              manual_adjust_flag=flag, revenue_forecast=100,
                              manual_revenue_month=0)
                result = calculate_final_values(values)
                self.assertEqual(value, result["final_revenue_segment"])
                self.assertIs(type(value), type(result["final_revenue_segment"]))
                self.assertEqual(Decimal("0.00"), result["final_revenue_forecast"])

    def test_month_normalization_and_visible_ambiguity(self):
        for mode in ("rpd", "cpd"):
            for ref, raw, expected in (
                ("2026-09", "10", "2026-10"),
                ("2026-12", "1月", "2027-01"),
                ("2026-01", "12", "2025-12"),
                ("2026-03", "9月", YEAR_REQUIRED_HINT),
                (None, "9月", YEAR_REQUIRED_HINT),
                (None, "Sep-26", "2026-09"),
                (None, "2026-13", INVALID_MONTH_HINT),
            ):
                with self.subTest(mode=mode, raw=raw, ref=ref):
                    result = calculate_final_values({
                        f"revenue_month_{mode}": ref,
                        f"manual_revenue_forecast_{mode}": raw,
                    })
                    self.assertEqual(expected, result[f"final_revenue_month_{mode}"])

    def test_first_run_formulas_metadata_styles_and_calculation_properties(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            path = root / "out.xlsx"
            _run(_write_sources(root, "first", variant="first"), path)
            wb = load_workbook(path)
            try:
                sheet = wb["基表"]
                config = load_config(CONFIG)
                ids = [c["id"] for c in config.base_columns]
                self.assertEqual(40, sheet.max_column)
                self.assertEqual("auto", wb.calculation.calcMode)
                self.assertTrue(wb.calculation.fullCalcOnLoad)
                self.assertTrue(wb.calculation.forceFullCalc)
                self.assertTrue(sheet.protection.sheet)
                self.assertFalse(sheet.protection.autoFilter)
                for manual, _ in FINAL_FIELD_SOURCES.values():
                    self.assertFalse(sheet.cell(2, ids.index(manual) + 1).protection.locked)
                for number in range(2, sheet.max_row + 1):
                    refs = {f: f"{get_column_letter(i)}{number}"
                            for i, f in enumerate(ids, 1)}
                    for final, formula in final_formulas(refs).items():
                        cell = sheet.cell(number, ids.index(final) + 1)
                        self.assertEqual(formula, cell.value)
                        self.assertEqual("f", cell.data_type)
                        self.assertLess(len(formula), 8192)
                        self.assertNotEqual("00FFF2CC", cell.fill.fgColor.rgb)
                        self.assertTrue(cell.protection.locked)
                        self.assertNotIn(refs["manual_adjust_flag"], formula)
                        if "month" in final:
                            self.assertEqual("@", cell.number_format)
                        elif final.endswith("forecast"):
                            self.assertEqual("#,##0.00", cell.number_format)
                self.assertEqual(0, len(sheet.data_validations.dataValidation))
                for s in wb:
                    if s.sheet_state == "visible":
                        self.assertEqual("A2", s.freeze_panes)
                        self.assertIsNotNone(s.auto_filter.ref)
                        self.assertFalse(s.tables)
                meta_ids = {r[0] for r in wb["_tool_meta"].iter_rows(values_only=True)}
                self.assertTrue(set(FINAL_FIELD_SOURCES) <= meta_ids)
            finally:
                wb.close()
            with ZipFile(path) as archive:
                ns = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
                tree = ET.fromstring(archive.read("xl/worksheets/sheet1.xml"))
                self.assertEqual(4 * (sheet.max_row - 1), len(tree.findall(".//x:f", ns)))
                self.assertIsNotNone(tree.find("x:autoFilter", ns))

    def test_previous_without_final_columns_and_new_without_caches(self):
        for format_kind in ("old", "no-cache", "stale-final-values"):
            with self.subTest(format_kind=format_kind), TemporaryDirectory() as tmp:
                root = Path(tmp)
                sources = _write_sources(root, "roundtrip", variant="first")
                previous, current = root / "previous.xlsx", root / "current.xlsx"
                _run(sources, previous)
                _set_manual_inputs(previous, "C001", "SC-A", segment_flag=False,
                                   rpd="2026-9", cpd="2026-10", amount=0)
                if format_kind == "old":
                    _remove_fields_from_result(previous, tuple(FINAL_FIELD_SOURCES))
                else:
                    _rename_metadata_backed_field(previous, "manual_revenue_segment", "用户分段")
                    _insert_noncritical_column_and_move_to_end(previous, "用户分段")
                    if format_kind == "stale-final-values":
                        wb = load_workbook(previous)
                        try:
                            names = {c["name"] for c in load_config(CONFIG).base_columns
                                     if c["id"] in FINAL_FIELD_SOURCES}
                            for cell in wb["基表"][1]:
                                if cell.value in names:
                                    for number in range(2, wb["基表"].max_row + 1):
                                        wb["基表"].cell(number, cell.column).value = "过时最终值"
                            wb.save(previous)
                        finally:
                            wb.close()
                issues = IssueLog()
                state = ExcelInputAdapter().read_previous(previous, load_config(CONFIG), issues)
                row = state.rows[("C001", "sc-a")].values
                self.assertTrue(set(FINAL_FIELD_SOURCES).isdisjoint(row))
                self.assertEqual(False, row["manual_revenue_segment"])
                self.assertEqual("2026-09", row["manual_revenue_forecast_rpd"])
                self.assertEqual(Decimal("0.00"), row["manual_revenue_month"])
                self.assertFalse(any(i.field in FINAL_FIELD_SOURCES for i in issues.items))
                result = _run(sources, current, previous=previous)
                self.assertEqual(0, result.rpd_change_count)
                self.assertEqual(0, result.cpd_change_count)
                wb = load_workbook(current)
                try:
                    data = _base_rows(wb["基表"])[("C001", "SC-A")]
                    self.assertIs(False, data["调整收入分段类别"])
                    self.assertEqual(0, data["调整金额"])
                    self.assertTrue(data["最终收入预测"].startswith("=IF("))
                finally:
                    wb.close()


class FormulaRecalculationTest(unittest.TestCase):
    def test_formula_values_match_python_after_edit_and_clear(self):
        office = shutil.which("libreoffice") or shutil.which("soffice")
        if not office:
            if os.environ.get("REQUIRE_FORMULA_ENGINE") == "1":
                self.fail("LibreOffice must be installed for the formula verification gate")
            self.skipTest("Independent formula engine available in Linux verification job")
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            config = load_config(CONFIG)
            cases = formula_cases()
            initial = [{k: v for k, v in case.items() if not k.startswith("manual_")}
                       for case in cases]
            rows = [BaseRow(dict(v, contract_no=f"T{i}", supply_center="深供",
                                 **calculate_final_values(v))) for i, v in enumerate(initial)]
            source = root / "edit.xlsx"
            ExcelOutputAdapter().write(source, rows, [], [], [], IssueLog(), config)
            indexes = {c["id"]: i for i, c in enumerate(config.base_columns, 1)}
            # Edit an already-generated workbook and then clear the same inputs.
            for stage in ("edited", "cleared"):
                if stage == "edited":
                    wb = load_workbook(source)
                    try:
                        for number, values in enumerate(cases, 2):
                            for field, value in values.items():
                                if field.startswith("manual_"):
                                    cell = wb["基表"].cell(number, indexes[field])
                                    cell.value = value
                                    # Calc coerces boolean inputs to numeric when
                                    # importing a numeric format. Keep their real
                                    # type in this cross-engine input fixture.
                                    if isinstance(value, bool):
                                        cell.number_format = "General"
                        wb.save(source)
                    finally:
                        wb.close()
                if stage == "cleared":
                    wb = load_workbook(source)
                    try:
                        for number in range(2, wb["基表"].max_row + 1):
                            for manual, _ in FINAL_FIELD_SOURCES.values():
                                wb["基表"].cell(number, indexes[manual]).value = None
                        wb.save(source)
                    finally:
                        wb.close()
                    cases = [{k: v for k, v in c.items() if not k.startswith("manual_")}
                             for c in cases]
                out = root / stage
                out.mkdir()
                subprocess.run([
                    office, f"-env:UserInstallation={(root / 'office-profile').as_uri()}",
                    "--headless", "--convert-to", "xlsx", "--outdir", str(out), str(source),
                ], check=True, capture_output=True, timeout=120)
                wb = load_workbook(out / source.name, data_only=True)
                try:
                    for number, values in enumerate(cases, 2):
                        for field, expected in calculate_final_values(values).items():
                            actual = wb["基表"].cell(number, indexes[field]).value
                            if isinstance(expected, Decimal):
                                expected = float(expected)
                            self.assertEqual(expected, actual, (stage, number, field, values))
                            if isinstance(expected, bool):
                                self.assertIs(expected, actual)
                finally:
                    wb.close()


if __name__ == "__main__":
    # Export only JSON for optional independent evaluator checks in development.
    from revenue_tool.adapters.final_revenue_formulas import final_formulas
    refs = {c["id"]: f"{get_column_letter(i)}2"
            for i, c in enumerate(load_config(CONFIG).base_columns, 1)}
    print(json.dumps({"refs": refs, "formulas": final_formulas(refs),
                      "cases": [{"inputs": v, "expected": calculate_final_values(v)}
                                for v in formula_cases()]}, ensure_ascii=False, default=float))
