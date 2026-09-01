from __future__ import annotations

from datetime import datetime
from pathlib import Path
from tempfile import TemporaryDirectory
import unittest

from openpyxl import load_workbook

from tests.test_pipeline import (
    CONFIG,
    EXPECTED_BASE_HEADERS,
    _base_rows,
    _rows_by_key,
    _run,
    _set_manual_values,
    _write_sources,
)


MANUAL_MONTH_HEADERS = (
    "调整月份（按RPD）",
    "调整月份（按CPD）",
)

POST_32_FIELD_IDS = (
    "revenue_forecast",
    "manual_revenue_segment_flag",
    "manual_revenue_forecast_rpd",
    "manual_revenue_forecast_cpd",
)

LEGACY_MANUAL_DISPLAY_NAMES = {
    "是否手工调整预测": "是否手工调整收入月份",
    "调整月份（按RPD）": "手工调整收入预测（按RPD）",
    "调整月份（按CPD）": "手工调整收入预测（按CPD）",
    "调整金额": "手工调整收入月份",
}


class ManualRevenueForecastTest(unittest.TestCase):
    def test_first_run_writes_computed_forecast_and_blank_manual_adjustments(
        self,
    ) -> None:
        with TemporaryDirectory() as temporary:
            directory = Path(temporary)
            sources = _write_sources(directory, "first", variant="first")
            output = directory / "result.xlsx"

            _run(sources, output)

            workbook = load_workbook(output)
            try:
                base = workbook["基表"]
                headers = [cell.value for cell in base[1]]
                self.assertEqual(36, len(headers))
                self.assertEqual(EXPECTED_BASE_HEADERS, headers)
                indexes = {cell.value: cell.column for cell in base[1]}
                for row_number in range(2, base.max_row + 1):
                    forecast_cell = base.cell(
                        row_number,
                        indexes["收入预测"],
                    )
                    self.assertEqual("#,##0.00", forecast_cell.number_format)
                    self.assertNotEqual("f", forecast_cell.data_type)
                    self.assertNotEqual(
                        "FFF2CC",
                        forecast_cell.fill.fgColor.rgb[-6:],
                    )
                    for header in MANUAL_MONTH_HEADERS:
                        cell = base.cell(row_number, indexes[header])
                        self.assertIsNone(cell.value)
                        self.assertEqual("@", cell.number_format)
                        self.assertEqual("solid", cell.fill.fill_type)
                        self.assertEqual("FFF2CC", cell.fill.fgColor.rgb[-6:])
                    for header in (
                        "是否修改收入分段类别",
                        "调整金额",
                    ):
                        cell = base.cell(row_number, indexes[header])
                        self.assertIsNone(cell.value)
                        self.assertEqual("solid", cell.fill.fill_type)
                        self.assertEqual("FFF2CC", cell.fill.fgColor.rgb[-6:])
                    self.assertEqual(
                        "#,##0.00",
                        base.cell(row_number, indexes["调整金额"]).number_format,
                    )
                    self.assertEqual(
                        "@",
                        base.cell(
                            row_number, indexes["收入年月（按RPD）"]
                        ).number_format,
                    )
                    self.assertEqual(
                        "@",
                        base.cell(
                            row_number, indexes["收入年月（按CPD）"]
                        ).number_format,
                    )

                rows = _base_rows(base)
                self.assertEqual(100, rows[("C001", "SC-A")]["遗留量"])
                self.assertEqual(50, rows[("C001", "SC-A")]["当月新订货"])
                self.assertEqual(150, rows[("C001", "SC-A")]["收入预测"])

                metadata = workbook["_tool_meta"]
                field_ids = []
                for row_number in range(5, metadata.max_row + 1):
                    value = metadata.cell(row_number, 1).value
                    if value is None:
                        break
                    field_ids.append(value)
                self.assertEqual(36, len(field_ids))
                for field in POST_32_FIELD_IDS:
                    self.assertIn(field, field_ids)
            finally:
                workbook.close()

    def test_manual_adjustments_inherit_and_previous_forecast_is_ignored(
        self,
    ) -> None:
        with TemporaryDirectory() as temporary:
            directory = Path(temporary)
            sources = _write_sources(directory, "roundtrip", variant="first")
            previous = directory / "previous.xlsx"
            output = directory / "current.xlsx"
            _run(sources, previous)
            _set_manual_inputs(
                previous,
                "C001",
                "SC-A",
                revenue_forecast=123.456,
                segment_flag="N",
                rpd="2026-09",
                cpd="2026-10",
                amount=0,
            )
            _set_manual_inputs(
                previous,
                "C001",
                "SC-B",
                rpd="2026-11",
                amount=-7.5,
            )

            _run(sources, output, previous=previous)

            workbook = load_workbook(output, data_only=True)
            try:
                rows = _base_rows(workbook["基表"])
                inherited = rows[("C001", "SC-A")]
                self.assertEqual(150, inherited["收入预测"])
                self.assertEqual("N", inherited["是否修改收入分段类别"])
                self.assertEqual("2026-09", inherited["调整月份（按RPD）"])
                self.assertEqual("2026-10", inherited["调整月份（按CPD）"])
                self.assertEqual(0, inherited["调整金额"])
                positive = rows[("C001", "SC-B")]
                self.assertEqual(150, positive["收入预测"])
                self.assertEqual(
                    "2026-11",
                    positive["调整月份（按RPD）"],
                )
                self.assertEqual(-7.5, positive["调整金额"])
                blank = rows[("C003", "SC-C")]
                for header in MANUAL_MONTH_HEADERS:
                    self.assertIsNone(blank[header])
            finally:
                workbook.close()

    def test_previous_revenue_forecast_is_ignored_without_manual_issue(
        self,
    ) -> None:
        with TemporaryDirectory() as temporary:
            directory = Path(temporary)
            sources = _write_sources(directory, "invalid", variant="first")
            previous = directory / "previous.xlsx"
            output = directory / "current.xlsx"
            _run(sources, previous)
            _set_manual_inputs(
                previous,
                "C001",
                "SC-A",
                revenue_forecast="not-an-amount",
            )

            _run(sources, output, previous=previous)

            workbook = load_workbook(output, data_only=True)
            try:
                row = _base_rows(workbook["基表"])[("C001", "SC-A")]
                self.assertEqual(150, row["收入预测"])
                matching = [
                    values
                    for values in workbook["异常清单"].iter_rows(
                        min_row=2,
                        values_only=True,
                    )
                    if values[0] == "INVALID_PREVIOUS_MANUAL_AMOUNT"
                    and values[6] == "revenue_forecast"
                ]
                self.assertEqual([], matching)
            finally:
                workbook.close()

    def test_invalid_previous_manual_adjustment_is_reported_as_blank(
        self,
    ) -> None:
        with TemporaryDirectory() as temporary:
            directory = Path(temporary)
            sources = _write_sources(
                directory,
                "invalid-adjustment",
                variant="first",
            )
            previous = directory / "previous.xlsx"
            output = directory / "current.xlsx"
            _run(sources, previous)
            _set_manual_inputs(
                previous,
                "C001",
                "SC-A",
                rpd="not-a-month",
            )

            _run(sources, output, previous=previous)

            workbook = load_workbook(output, data_only=True)
            try:
                row = _base_rows(workbook["基表"])[("C001", "SC-A")]
                self.assertIsNone(row["调整月份（按RPD）"])
                matching = [
                    values
                    for values in workbook["异常清单"].iter_rows(
                        min_row=2,
                        values_only=True,
                    )
                    if values[0] == "INVALID_PREVIOUS_MANUAL_MONTH"
                    and values[6] == "manual_revenue_forecast_rpd"
                ]
                self.assertEqual(1, len(matching))
                self.assertEqual("not-a-month", matching[0][7])
            finally:
                workbook.close()

    def test_invalid_previous_manual_amount_and_flag_are_reported_as_blank(
        self,
    ) -> None:
        with TemporaryDirectory() as temporary:
            directory = Path(temporary)
            sources = _write_sources(directory, "invalid-inputs", variant="first")
            previous = directory / "previous.xlsx"
            output = directory / "current.xlsx"
            _run(sources, previous)
            _set_manual_inputs(
                previous,
                "C001",
                "SC-A",
                segment_flag="MAYBE",
                amount="not-an-amount",
            )

            _run(sources, output, previous=previous)

            workbook = load_workbook(output, data_only=True)
            try:
                row = _base_rows(workbook["基表"])[("C001", "SC-A")]
                self.assertIsNone(row["是否修改收入分段类别"])
                self.assertIsNone(row["调整金额"])
                matching = {
                    (values[0], values[6])
                    for values in workbook["异常清单"].iter_rows(
                        min_row=2,
                        values_only=True,
                    )
                }
                self.assertIn(
                    (
                        "INVALID_PREVIOUS_MANUAL_FLAG",
                        "manual_revenue_segment_flag",
                    ),
                    matching,
                )
                self.assertIn(
                    (
                        "INVALID_PREVIOUS_MANUAL_AMOUNT",
                        "manual_revenue_month",
                    ),
                    matching,
                )
            finally:
                workbook.close()

    def test_date_cells_for_manual_months_are_normalized_to_text(self) -> None:
        with TemporaryDirectory() as temporary:
            directory = Path(temporary)
            sources = _write_sources(directory, "date-months", variant="first")
            previous = directory / "previous.xlsx"
            output = directory / "current.xlsx"
            _run(sources, previous)
            _set_manual_inputs(
                previous,
                "C001",
                "SC-A",
                rpd=datetime(2026, 9, 15),
                cpd=datetime(2026, 10, 1),
            )

            _run(sources, output, previous=previous)

            workbook = load_workbook(output)
            try:
                base = workbook["基表"]
                rows = _base_rows(base)
                row = rows[("C001", "SC-A")]
                self.assertEqual("2026-09", row["调整月份（按RPD）"])
                self.assertEqual("2026-10", row["调整月份（按CPD）"])
                headers = {cell.value: cell.column for cell in base[1]}
                row_number = next(
                    number
                    for number in range(2, base.max_row + 1)
                    if base.cell(number, headers["合同号"]).value == "C001"
                    and base.cell(number, headers["履行供应中心"]).value
                    == "SC-A"
                )
                self.assertEqual(
                    "@",
                    base.cell(
                        row_number, headers["调整月份（按RPD）"]
                    ).number_format,
                )
            finally:
                workbook.close()

    def test_short_months_roundtrip_after_column_move_with_metadata_or_alias(
        self,
    ) -> None:
        for metadata_mode in ("stable-id", "legacy-alias"):
            with self.subTest(metadata_mode=metadata_mode):
                with TemporaryDirectory() as temporary:
                    directory = Path(temporary)
                    sources = _write_sources(
                        directory,
                        f"short-{metadata_mode}",
                        variant="first",
                    )
                    previous = directory / "previous.xlsx"
                    output = directory / "current.xlsx"
                    _run(sources, previous)
                    _set_previous_month_context(
                        previous,
                        "C001",
                        "SC-A",
                        automatic_rpd="2026-06",
                        automatic_cpd="2026-07",
                        manual_rpd="9月",
                        manual_cpd=10,
                    )
                    if metadata_mode == "stable-id":
                        _rename_metadata_backed_field(
                            previous,
                            "manual_revenue_forecast_rpd",
                            "用户填写RPD月份",
                        )
                        moved_header = "用户填写RPD月份"
                    else:
                        _rename_manual_headers_to_legacy(
                            previous,
                            keep_metadata=False,
                        )
                        moved_header = "手工调整收入预测（按RPD）"
                    _insert_noncritical_column_and_move_to_end(
                        previous,
                        moved_header,
                    )

                    _run(sources, output, previous=previous)

                    workbook = load_workbook(output)
                    try:
                        base = workbook["基表"]
                        row = _base_rows(base)[("C001", "SC-A")]
                        self.assertEqual(
                            "2026-09", row["调整月份（按RPD）"]
                        )
                        self.assertEqual(
                            "2026-10", row["调整月份（按CPD）"]
                        )
                        headers = {
                            cell.value: cell.column for cell in base[1]
                        }
                        row_number = next(
                            number
                            for number in range(2, base.max_row + 1)
                            if base.cell(number, headers["合同号"]).value
                            == "C001"
                            and base.cell(
                                number, headers["履行供应中心"]
                            ).value
                            == "SC-A"
                        )
                        for header in MANUAL_MONTH_HEADERS:
                            cell = base.cell(row_number, headers[header])
                            self.assertEqual("@", cell.number_format)
                            self.assertNotIsInstance(
                                cell.value, (datetime,)
                            )
                        issue_codes = {
                            values[0]
                            for values in workbook["异常清单"].iter_rows(
                                min_row=2,
                                values_only=True,
                            )
                            if values[6]
                            in {
                                "manual_revenue_forecast_rpd",
                                "manual_revenue_forecast_cpd",
                            }
                        }
                        self.assertEqual(set(), issue_codes)
                    finally:
                        workbook.close()

    def test_month_only_ambiguity_and_missing_reference_require_year(
        self,
    ) -> None:
        with TemporaryDirectory() as temporary:
            directory = Path(temporary)
            sources = _write_sources(directory, "year-required", variant="first")
            previous = directory / "previous.xlsx"
            output = directory / "current.xlsx"
            _run(sources, previous)
            _set_previous_month_context(
                previous,
                "C001",
                "SC-A",
                automatic_rpd="2026-03",
                automatic_cpd="2026-03",
                manual_rpd="9月",
            )
            _set_previous_month_context(
                previous,
                "C002",
                None,
                automatic_rpd=None,
                automatic_cpd=None,
                manual_cpd="10",
            )

            _run(sources, output, previous=previous)

            workbook = load_workbook(output, data_only=True)
            try:
                rows = _base_rows(workbook["基表"])
                self.assertIsNone(
                    rows[("C001", "SC-A")]["调整月份（按RPD）"]
                )
                self.assertIsNone(rows[("C002", None)]["调整月份（按CPD）"])
                matching = [
                    values
                    for values in workbook["异常清单"].iter_rows(
                        min_row=2,
                        values_only=True,
                    )
                    if values[0] == "MANUAL_MONTH_YEAR_REQUIRED"
                ]
                self.assertEqual(2, len(matching))
                self.assertEqual(
                    {
                        (
                            "C001 | SC-A",
                            "manual_revenue_forecast_rpd",
                            "9月",
                        ),
                        (
                            "C002 | ",
                            "manual_revenue_forecast_cpd",
                            "10",
                        ),
                    },
                    {(row[5], row[6], str(row[7])) for row in matching},
                )
                self.assertTrue(
                    all("请填写完整年月" in row[8] for row in matching)
                )
            finally:
                workbook.close()

    def test_business_blank_markers_remain_blank_without_manual_issues(
        self,
    ) -> None:
        with TemporaryDirectory() as temporary:
            directory = Path(temporary)
            sources = _write_sources(directory, "blank-markers", variant="first")
            previous = directory / "previous.xlsx"
            output = directory / "current.xlsx"
            _run(sources, previous)
            _set_manual_inputs(
                previous,
                "C001",
                "SC-A",
                segment_flag="VALUE",
                rpd="VALUE",
                cpd="(空白)",
                amount="#VALUE!",
            )

            _run(sources, output, previous=previous)

            workbook = load_workbook(output, data_only=True)
            try:
                row = _base_rows(workbook["基表"])[("C001", "SC-A")]
                for header in (
                    "是否修改收入分段类别",
                    "调整月份（按RPD）",
                    "调整月份（按CPD）",
                    "调整金额",
                ):
                    self.assertIsNone(row[header])
                manual_issue_codes = {
                    values[0]
                    for values in workbook["异常清单"].iter_rows(
                        min_row=2,
                        values_only=True,
                    )
                    if values[6]
                    in {
                        "manual_revenue_segment_flag",
                        "manual_revenue_forecast_rpd",
                        "manual_revenue_forecast_cpd",
                        "manual_revenue_month",
                    }
                }
                self.assertEqual(set(), manual_issue_codes)
            finally:
                workbook.close()

    def test_old_35_column_result_keeps_old_manual_values_and_new_flag_blank(
        self,
    ) -> None:
        with TemporaryDirectory() as temporary:
            directory = Path(temporary)
            sources = _write_sources(directory, "old-35", variant="first")
            previous = directory / "previous.xlsx"
            output = directory / "current.xlsx"
            _run(sources, previous)
            _set_manual_values(previous, "C001", "SC-A")
            _remove_fields_from_result(
                previous,
                ("manual_revenue_segment_flag",),
            )

            _run(sources, output, previous=previous)

            workbook = load_workbook(output, data_only=True)
            try:
                row = _base_rows(workbook["基表"])[("C001", "SC-A")]
                self.assertIsNone(row["是否修改收入分段类别"])
                self.assertEqual("2026-04", row["调整月份（按RPD）"])
                self.assertEqual("2026-05", row["调整月份（按CPD）"])
                self.assertEqual(125.5, row["调整金额"])
                self.assertEqual("业务确认", row["调整备注"])
            finally:
                workbook.close()

    def test_old_32_column_result_remains_usable(self) -> None:
        with TemporaryDirectory() as temporary:
            directory = Path(temporary)
            first_sources = _write_sources(directory, "first", variant="first")
            second_sources = _write_sources(directory, "second", variant="second")
            previous = directory / "previous-32-columns.xlsx"
            output = directory / "current.xlsx"
            _run(first_sources, previous)
            _set_manual_values(previous, "C001", "SC-A")
            _remove_fields_from_result(previous, POST_32_FIELD_IDS)

            result = _run(second_sources, output, previous=previous)

            self.assertGreaterEqual(result.rpd_change_count, 1)
            self.assertGreaterEqual(result.cpd_change_count, 1)
            workbook = load_workbook(output, data_only=True)
            try:
                inherited = _base_rows(workbook["基表"])[("C001", "SC-A")]
                self.assertEqual("Y", inherited["是否手工调整预测"])
                self.assertEqual(125.5, inherited["调整金额"])
                self.assertEqual("业务确认", inherited["调整备注"])
                self.assertEqual(
                    inherited["遗留量"] + inherited["当月新订货"],
                    inherited["收入预测"],
                )
                self.assertIsNone(inherited["是否修改收入分段类别"])
                for header in MANUAL_MONTH_HEADERS:
                    self.assertIsNone(inherited[header])
                codes = {
                    values[0]
                    for values in workbook["异常清单"].iter_rows(
                        min_row=2,
                        values_only=True,
                    )
                }
                self.assertNotIn("PREVIOUS_BASE_UNUSABLE", codes)
            finally:
                workbook.close()

    def test_legacy_display_names_inherit_with_or_without_metadata(self) -> None:
        for keep_metadata in (True, False):
            with self.subTest(keep_metadata=keep_metadata):
                with TemporaryDirectory() as temporary:
                    directory = Path(temporary)
                    sources = _write_sources(
                        directory,
                        "legacy-names",
                        variant="first",
                    )
                    previous = directory / "previous.xlsx"
                    output = directory / "current.xlsx"
                    _run(sources, previous)
                    _set_manual_values(previous, "C001", "SC-A")
                    _rename_manual_headers_to_legacy(
                        previous,
                        keep_metadata=keep_metadata,
                    )

                    _run(sources, output, previous=previous)

                    workbook = load_workbook(output, data_only=True)
                    try:
                        inherited = _base_rows(workbook["基表"])[
                            ("C001", "SC-A")
                        ]
                        self.assertEqual(
                            "Y", inherited["是否修改收入分段类别"]
                        )
                        self.assertEqual("Y", inherited["是否手工调整预测"])
                        self.assertEqual(
                            "2026-04", inherited["调整月份（按RPD）"]
                        )
                        self.assertEqual(
                            "2026-05", inherited["调整月份（按CPD）"]
                        )
                        self.assertEqual(125.5, inherited["调整金额"])
                        self.assertEqual("业务确认", inherited["调整备注"])
                        unavailable = {
                            values[6]
                            for values in workbook["异常清单"].iter_rows(
                                min_row=2,
                                values_only=True,
                            )
                            if values[0] == "PREVIOUS_FIELD_UNAVAILABLE"
                        }
                        self.assertTrue(
                            {
                                "manual_adjust_flag",
                                "manual_revenue_forecast_rpd",
                                "manual_revenue_forecast_cpd",
                                "manual_revenue_month",
                            }.isdisjoint(unavailable)
                        )
                    finally:
                        workbook.close()

    def test_deep_supply_keeps_contract_amounts_and_other_center_displays_zero(
        self,
    ) -> None:
        with TemporaryDirectory() as temporary:
            directory = Path(temporary)
            sources = _write_sources(directory, "deep", variant="first")
            _replace_demand_center(sources[2], "C001", "SC-A", "　深供　")
            _replace_transit_center(sources[3], "SC-A", "深供")
            _set_demand_value(sources[2], "C001", "SC-B", "天_ATA", None)
            _append_monthly_contract(sources[1], "C002", 22)
            output = directory / "result.xlsx"

            _run(sources, output)

            workbook = load_workbook(output, data_only=True)
            try:
                rows = _base_rows(workbook["基表"])
                deep = rows[("C001", "深供")]
                other = rows[("C001", "SC-B")]
                self.assertEqual(100, deep["遗留量"])
                self.assertEqual(50, deep["当月新订货"])
                self.assertEqual(150, deep["收入预测"])
                self.assertEqual(0, other["遗留量"])
                self.assertEqual(0, other["当月新订货"])
                self.assertEqual(0, other["收入预测"])
                self.assertEqual("订未发", other["收入分段类别"])
                self.assertEqual(300, rows[("C003", "SC-C")]["遗留量"])
                self.assertEqual(22, rows[("C002", None)]["当月新订货"])
                self.assertEqual(22, rows[("C002", None)]["收入预测"])
                self.assertEqual(
                    150,
                    sum(
                        row["收入预测"]
                        for key, row in rows.items()
                        if key[0] == "C001"
                    ),
                )
                supply_rows = _rows_by_key(
                    workbook["供应需要提拉诉求清单粗表"]
                )
                self.assertEqual(100, supply_rows[("C001", "深供")]["遗留量"])
                self.assertEqual(
                    50,
                    supply_rows[("C001", "深供")]["当月新订货"],
                )

                c001_missing_deep = [
                    values
                    for values in workbook["异常清单"].iter_rows(
                        min_row=2,
                        values_only=True,
                    )
                    if values[0] == "MULTI_CENTER_DEEP_SUPPLY_NOT_FOUND"
                    and values[5] == "C001"
                ]
                self.assertEqual([], c001_missing_deep)
            finally:
                workbook.close()

    def test_multi_center_without_deep_keeps_all_amounts_and_reports_issue(
        self,
    ) -> None:
        with TemporaryDirectory() as temporary:
            directory = Path(temporary)
            sources = _write_sources(directory, "no-deep", variant="first")
            output = directory / "result.xlsx"

            _run(sources, output)

            workbook = load_workbook(output, data_only=True)
            try:
                rows = _base_rows(workbook["基表"])
                for center in ("SC-A", "SC-B"):
                    self.assertEqual(100, rows[("C001", center)]["遗留量"])
                    self.assertEqual(50, rows[("C001", center)]["当月新订货"])
                    self.assertEqual(150, rows[("C001", center)]["收入预测"])
                matching = [
                    values
                    for values in workbook["异常清单"].iter_rows(
                        min_row=2,
                        values_only=True,
                    )
                    if values[0] == "MULTI_CENTER_DEEP_SUPPLY_NOT_FOUND"
                    and values[5] == "C001"
                ]
                self.assertEqual(1, len(matching))
                self.assertIn("SC-A", matching[0][7])
                self.assertIn("SC-B", matching[0][7])
            finally:
                workbook.close()

    def test_control_flag_mismatch_is_cancelled_and_stock_flag_still_rules(
        self,
    ) -> None:
        with TemporaryDirectory() as temporary:
            directory = Path(temporary)
            sources = _write_sources(directory, "flags", variant="first")
            _set_all_demand_values(
                sources[2],
                "C003",
                {
                    "备货总控标识": "N",
                    "发货总控标识": "Y",
                },
            )
            output = directory / "result.xlsx"

            _run(sources, output)

            workbook = load_workbook(output, data_only=True)
            try:
                rows = _base_rows(workbook["基表"])
                self.assertEqual("未解锁", rows[("C001", "SC-B")]["是否解锁备货"])
                self.assertEqual("已解锁", rows[("C003", "SC-C")]["是否解锁备货"])
                codes = {
                    values[0]
                    for values in workbook["异常清单"].iter_rows(
                        min_row=2,
                        values_only=True,
                    )
                }
                self.assertNotIn("CONTROL_FLAG_MISMATCH", codes)
            finally:
                workbook.close()

    def test_invalid_shipment_flag_remains_a_field_validation_issue(
        self,
    ) -> None:
        with TemporaryDirectory() as temporary:
            directory = Path(temporary)
            sources = _write_sources(
                directory,
                "shipment-invalid",
                variant="first",
            )
            _set_demand_value(
                sources[2],
                "C003",
                "SC-C",
                "发货总控标识",
                "X",
            )
            output = directory / "result.xlsx"

            _run(sources, output)

            workbook = load_workbook(output, data_only=True)
            try:
                c003 = _base_rows(workbook["基表"])[("C003", "SC-C")]
                self.assertEqual("未解锁", c003["是否解锁备货"])
                matching = [
                    row
                    for row in workbook["异常清单"].iter_rows(
                        min_row=2,
                        values_only=True,
                    )
                    if row[0] == "INVALID_ENUM_VALUE"
                    and row[6] == "shipment_control_flag"
                ]
                self.assertGreaterEqual(len(matching), 1)
                codes = {
                    row[0]
                    for row in workbook["异常清单"].iter_rows(
                        min_row=2,
                        values_only=True,
                    )
                }
                self.assertNotIn("CONTROL_FLAG_MISMATCH", codes)
            finally:
                workbook.close()


def _set_manual_inputs(
    path: Path,
    contract: str,
    center: str | None,
    *,
    revenue_forecast=None,
    segment_flag=None,
    rpd=None,
    cpd=None,
    amount=None,
) -> None:
    workbook = load_workbook(path)
    try:
        sheet = workbook["基表"]
        headers = {cell.value: cell.column for cell in sheet[1]}
        for row_number in range(2, sheet.max_row + 1):
            if (
                sheet.cell(row_number, headers["合同号"]).value == contract
                and sheet.cell(row_number, headers["履行供应中心"]).value == center
            ):
                sheet.cell(row_number, headers["收入预测"]).value = revenue_forecast
                sheet.cell(
                    row_number,
                    headers["是否修改收入分段类别"],
                ).value = segment_flag
                sheet.cell(
                    row_number,
                    headers["调整月份（按RPD）"],
                ).value = rpd
                sheet.cell(
                    row_number,
                    headers["调整月份（按CPD）"],
                ).value = cpd
                sheet.cell(
                    row_number,
                    headers["调整金额"],
                ).value = amount
                break
        workbook.save(path)
    finally:
        workbook.close()


def _set_previous_month_context(
    path: Path,
    contract: str,
    center: str | None,
    *,
    automatic_rpd,
    automatic_cpd,
    manual_rpd=None,
    manual_cpd=None,
) -> None:
    workbook = load_workbook(path)
    try:
        sheet = workbook["基表"]
        headers = {cell.value: cell.column for cell in sheet[1]}
        for row_number in range(2, sheet.max_row + 1):
            if (
                sheet.cell(row_number, headers["合同号"]).value == contract
                and sheet.cell(row_number, headers["履行供应中心"]).value
                == center
            ):
                sheet.cell(
                    row_number, headers["收入年月（按RPD）"]
                ).value = automatic_rpd
                sheet.cell(
                    row_number, headers["收入年月（按CPD）"]
                ).value = automatic_cpd
                sheet.cell(
                    row_number, headers["调整月份（按RPD）"]
                ).value = manual_rpd
                sheet.cell(
                    row_number, headers["调整月份（按CPD）"]
                ).value = manual_cpd
                break
        workbook.save(path)
    finally:
        workbook.close()


def _rename_metadata_backed_field(
    path: Path,
    field_id: str,
    display_name: str,
) -> None:
    workbook = load_workbook(path)
    try:
        metadata = workbook["_tool_meta"]
        old_display_name = None
        for row_number in range(5, metadata.max_row + 1):
            if metadata.cell(row_number, 1).value == field_id:
                old_display_name = metadata.cell(row_number, 2).value
                metadata.cell(row_number, 2).value = display_name
                break
        if old_display_name is None:
            raise AssertionError(f"metadata field not found: {field_id}")
        base = workbook["基表"]
        for cell in base[1]:
            if cell.value == old_display_name:
                cell.value = display_name
                break
        workbook.save(path)
    finally:
        workbook.close()


def _insert_noncritical_column_and_move_to_end(
    path: Path,
    header: str,
) -> None:
    workbook = load_workbook(path)
    try:
        sheet = workbook["基表"]
        sheet.insert_cols(3)
        sheet.cell(1, 3).value = "用户附加说明"
        source_column = next(
            cell.column for cell in sheet[1] if cell.value == header
        )
        values = [
            sheet.cell(row_number, source_column).value
            for row_number in range(1, sheet.max_row + 1)
        ]
        sheet.delete_cols(source_column)
        destination = sheet.max_column + 1
        for row_number, value in enumerate(values, start=1):
            sheet.cell(row_number, destination).value = value
        workbook.save(path)
    finally:
        workbook.close()


def _rename_manual_headers_to_legacy(
    path: Path,
    *,
    keep_metadata: bool,
) -> None:
    workbook = load_workbook(path)
    try:
        base = workbook["基表"]
        for cell in base[1]:
            if cell.value in LEGACY_MANUAL_DISPLAY_NAMES:
                cell.value = LEGACY_MANUAL_DISPLAY_NAMES[cell.value]
        if keep_metadata:
            metadata = workbook["_tool_meta"]
            for row_number in range(5, metadata.max_row + 1):
                value = metadata.cell(row_number, 2).value
                if value in LEGACY_MANUAL_DISPLAY_NAMES:
                    metadata.cell(row_number, 2).value = (
                        LEGACY_MANUAL_DISPLAY_NAMES[value]
                    )
        else:
            workbook.remove(workbook["_tool_meta"])
        workbook.save(path)
    finally:
        workbook.close()


def _remove_fields_from_result(path: Path, field_ids: tuple[str, ...]) -> None:
    workbook = load_workbook(path)
    try:
        base = workbook["基表"]
        metadata = workbook["_tool_meta"]
        names_by_id = {
            metadata.cell(row_number, 1).value: metadata.cell(row_number, 2).value
            for row_number in range(5, metadata.max_row + 1)
            if metadata.cell(row_number, 1).value in field_ids
        }
        header_indexes = {
            cell.value: cell.column
            for cell in base[1]
        }
        for field_id in sorted(
            field_ids,
            key=lambda item: header_indexes[names_by_id[item]],
            reverse=True,
        ):
            base.delete_cols(header_indexes[names_by_id[field_id]], 1)
        metadata_rows = [
            row_number
            for row_number in range(5, metadata.max_row + 1)
            if metadata.cell(row_number, 1).value in field_ids
        ]
        for row_number in reversed(metadata_rows):
            metadata.delete_rows(row_number, 1)
        workbook.save(path)
    finally:
        workbook.close()


def _replace_demand_center(
    path: Path,
    contract: str,
    old_center: str,
    new_center: str,
) -> None:
    workbook = load_workbook(path)
    try:
        sheet = workbook["要货明细"]
        headers = {cell.value: cell.column for cell in sheet[2]}
        for row_number in range(3, sheet.max_row + 1):
            if (
                sheet.cell(row_number, headers["原合同号"]).value == contract
                and sheet.cell(row_number, headers["供应中心简称"]).value
                == old_center
            ):
                sheet.cell(row_number, headers["供应中心简称"]).value = new_center
        workbook.save(path)
    finally:
        workbook.close()


def _replace_transit_center(path: Path, old_center: str, new_center: str) -> None:
    workbook = load_workbook(path)
    try:
        sheet = workbook["国家运输周期"]
        headers = {cell.value: cell.column for cell in sheet[2]}
        for row_number in range(3, sheet.max_row + 1):
            if sheet.cell(row_number, headers["供应中心"]).value == old_center:
                sheet.cell(row_number, headers["供应中心"]).value = new_center
        workbook.save(path)
    finally:
        workbook.close()


def _set_demand_value(
    path: Path,
    contract: str,
    center: str,
    header: str,
    value,
) -> None:
    workbook = load_workbook(path)
    try:
        sheet = workbook["要货明细"]
        headers = {cell.value: cell.column for cell in sheet[2]}
        for row_number in range(3, sheet.max_row + 1):
            if (
                sheet.cell(row_number, headers["原合同号"]).value == contract
                and sheet.cell(row_number, headers["供应中心简称"]).value == center
            ):
                sheet.cell(row_number, headers[header]).value = value
        workbook.save(path)
    finally:
        workbook.close()


def _set_all_demand_values(
    path: Path,
    contract: str,
    changes: dict[str, object],
) -> None:
    workbook = load_workbook(path)
    try:
        sheet = workbook["要货明细"]
        headers = {cell.value: cell.column for cell in sheet[2]}
        for row_number in range(3, sheet.max_row + 1):
            if sheet.cell(row_number, headers["原合同号"]).value != contract:
                continue
            for header, value in changes.items():
                sheet.cell(row_number, headers[header]).value = value
        workbook.save(path)
    finally:
        workbook.close()


def _append_monthly_contract(path: Path, contract: str, amount: int) -> None:
    workbook = load_workbook(path)
    try:
        sheet = workbook["当月订货"]
        sheet.append([contract, "地区M", "中国", "项目M", "BG-M", amount])
        workbook.save(path)
    finally:
        workbook.close()


if __name__ == "__main__":
    unittest.main()
