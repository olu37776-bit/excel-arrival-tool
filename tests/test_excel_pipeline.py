from datetime import date
from pathlib import Path
from tempfile import TemporaryDirectory
import unittest

from openpyxl import Workbook, load_workbook

from revenue_tool.application.pipeline import run_pipeline


ROOT = Path(__file__).resolve().parents[1]


class ExcelPipelineTest(unittest.TestCase):
    def test_pipeline_generates_revenue_output_and_month_delay(self) -> None:
        with TemporaryDirectory() as temporary:
            directory = Path(temporary)
            first_input = directory / "first.xlsx"
            second_input = directory / "second.xlsx"
            first_output = directory / "first-result.xlsx"
            second_output = directory / "second-result.xlsx"
            self._write_input(first_input, date(2026, 7, 1))
            self._write_input(second_input, date(2026, 8, 1))

            first = run_pipeline(first_input, first_output, ROOT / "config")
            second = run_pipeline(
                second_input,
                second_output,
                ROOT / "config",
                previous_path=first_output,
            )

            self.assertEqual(1, first.revenue_detail_count)
            self.assertEqual(1, second.delayed_count)
            workbook = load_workbook(second_output, data_only=True)
            try:
                self.assertEqual(
                    ["Revenue Summary", "Revenue Detail", "Comparison", "Run Info"],
                    workbook.sheetnames,
                )
                comparison = workbook["Comparison"]
                self.assertEqual("Delayed", comparison["I2"].value)
                self.assertEqual(1, comparison["H2"].value)
            finally:
                workbook.close()

    @staticmethod
    def _write_input(path: Path, plan_date: date) -> None:
        workbook = Workbook()
        prd = workbook.active
        prd.title = "PRD"
        prd.append(
            [
                "PO Number",
                "PRD",
                "Original PO Quantity",
                "Contract Number",
                "Shipping Point",
            ]
        )
        prd.append(["PO-1", date(2026, 1, 10), 100, "C-1", "SP-1"])
        prd.append(["PO-1", date(2026, 1, 5), 120, "C-1", "SP-1"])

        shipment = workbook.create_sheet("Shipment")
        shipment.append(
            [
                "PO Number",
                "Plan Month",
                "Plan Quantity",
                "Contract Number",
                "Shipping Point",
                "Trade Type",
                "Shipment ID",
                "Revenue Amount",
            ]
        )
        shipment.append(["PO-1", plan_date, 100, "C-1", "SP-1", "SEA", "S-1", 5000])
        shipment.append(["PO-1", plan_date, 100, "C-1", "SP-1", "SEA", "S-1", 5000])

        transit = workbook.create_sheet("Transit Days")
        transit.append(["Trade Type", "Transit Days"])
        transit.append(["SEA", 0])
        workbook.save(path)
        workbook.close()


if __name__ == "__main__":
    unittest.main()

