from __future__ import annotations

from datetime import date
from decimal import Decimal
from pathlib import Path
from tempfile import TemporaryDirectory
import unittest

from openpyxl import load_workbook

from revenue_tool.adapters.excel_reader import ExcelInputAdapter
from revenue_tool.adapters.previous_result_reader import PreviousResultReader
from revenue_tool.config import load_config
from revenue_tool.domain.models import (
    BaseRow,
    CONTRACT_ONLY_NO_DEMAND,
    DEMAND_CENTER,
    IssueLog,
    ParsedRow,
    PreviousData,
    SourceData,
)
from revenue_tool.services.calculation import RevenueEngine, _revenue_segment
from revenue_tool.services.normalization import business_key_identity
from tests.test_pipeline import (
    CONFIG,
    _base_rows,
    _run,
    _set_manual_values,
    _write_sources,
)


class NoDemandPlaceholderTest(unittest.TestCase):
    def test_row_kind_has_priority_over_dates_flags_and_amounts(self) -> None:
        self.assertEqual(
            "不要货",
            _revenue_segment(
                CONTRACT_ONLY_NO_DEMAND,
                "Y",
                date(2026, 1, 1),
                date(2026, 1, 2),
                date(2026, 1, 3),
                date(2026, 1, 4),
                Decimal("100.00"),
                Decimal("200.00"),
            ),
        )

    def test_domain_builds_one_explicit_contract_only_row(self) -> None:
        config = load_config(CONFIG)
        previous = PreviousData(
            {
                business_key_identity("C100", None): BaseRow(
                    {
                        "manual_adjust_flag": "Y",
                        "manual_revenue_month": "2026-09",
                        "adjustment_note": "保留",
                    },
                    row_kind=CONTRACT_ONLY_NO_DEMAND,
                )
            }
        )
        source = SourceData(
            {
                "legacy": Path("legacy.xlsx"),
                "monthly_order": Path("monthly.xlsx"),
                "demand_detail": Path("demand.xlsx"),
                "transit": Path("transit.xlsx"),
            },
            {
                "legacy": [
                    _row(
                        "legacy",
                        {
                            "contract_no": "C100",
                            "legacy_amount": Decimal("10.00"),
                            "bg": "BG-L",
                            "region": "地区L",
                            "country": "沙特阿拉伯",
                            "customer_group": "客户L",
                            "project_name": "项目L",
                        },
                    )
                ],
                "monthly_order": [
                    _row(
                        "monthly_order",
                        {
                            "contract_no": "C100",
                            "monthly_new_order": Decimal("20.00"),
                            "bg": "BG-M",
                        },
                    )
                ],
                "demand_detail": [],
                "transit": [],
            },
            {
                "legacy": "Sheet1",
                "monthly_order": "Sheet1",
                "demand_detail": "Sheet1",
                "transit": "Sheet1",
            },
        )
        issues = IssueLog()

        rows = RevenueEngine().calculate(source, previous, config, issues)

        self.assertEqual(1, len(rows))
        row = rows[0]
        self.assertEqual(CONTRACT_ONLY_NO_DEMAND, row.row_kind)
        self.assertEqual("C100", row.values["contract_no"])
        self.assertEqual(Decimal("10.00"), row.values["legacy_amount"])
        self.assertEqual(
            Decimal("20.00"), row.values["monthly_new_order"]
        )
        self.assertEqual("交付类", row.values["carryover_type"])
        self.assertIsNone(row.values["supply_center"])
        self.assertEqual("不要货", row.values["revenue_segment"])
        for field in (
            "multiple_supply_centers",
            "split_shipment",
            "multiple_demand",
            "split_supply",
        ):
            self.assertEqual("N", row.values[field])
        for field in (
            "transit_days",
            "ata",
            "asd",
            "rpd",
            "latest_asd",
            "latest_rpd",
            "cpd",
            "arrival_date_rpd",
            "arrival_date_cpd",
            "revenue_month_rpd",
            "revenue_month_cpd",
        ):
            self.assertIsNone(row.values[field])
        self.assertEqual("Y", row.values["manual_adjust_flag"])
        self.assertEqual([], issues.items)

    def test_monthly_only_and_both_sources_each_create_one_row(self) -> None:
        with TemporaryDirectory() as temporary:
            directory = Path(temporary)
            sources = _write_sources(directory, "contracts", variant="first")
            _append_monthly_contract(sources[1], "C002", 22)
            _append_monthly_contract(sources[1], "C008", 88)
            output = directory / "result.xlsx"

            _run(sources, output)

            workbook = load_workbook(output, data_only=True)
            try:
                rows = _contract_rows(workbook["合同收入预测"])
                c002 = rows["C002"]
                c008 = rows["C008"]
                self.assertEqual(22, c002["当月新订货"])
                self.assertEqual(88, c008["当月新订货"])
                self.assertEqual(0, c008["遗留量"])
                self.assertEqual("不要货", c008["分配状态"])
                self.assertEqual(0, c008["分配候选数"])
                codes = {
                    row[0].value
                    for row in workbook["异常清单"].iter_rows(min_row=2)
                }
                self.assertNotIn(
                    "CONTRACT_NOT_FOUND_IN_DEMAND_DETAIL", codes
                )
            finally:
                workbook.close()

    def test_existing_demand_without_center_is_not_no_demand(self) -> None:
        with TemporaryDirectory() as temporary:
            directory = Path(temporary)
            sources = _write_sources(directory, "missing-center", variant="first")
            _clear_all_supply_centers(sources[2], "C003")
            output = directory / "result.xlsx"

            _run(sources, output)

            workbook = load_workbook(output, data_only=True)
            try:
                rows = _base_rows(workbook["收入分配"])
                self.assertFalse(
                    any(contract == "C003" for contract, _center in rows)
                )
                issues = [
                    row
                    for row in workbook["异常清单"].iter_rows(
                        min_row=2, values_only=True
                    )
                    if row[0] == "MISSING_SUPPLY_CENTER"
                ]
                self.assertEqual(2, len(issues))
            finally:
                workbook.close()

    def test_placeholder_manual_values_do_not_leak_to_future_center(self) -> None:
        with TemporaryDirectory() as temporary:
            directory = Path(temporary)
            first_sources = _write_sources(directory, "first", variant="first")
            first_result = directory / "first.xlsx"
            second_result = directory / "second.xlsx"
            third_result = directory / "third.xlsx"
            _run(first_sources, first_result)
            _run(first_sources, second_result, previous=first_result)
            workbook = load_workbook(second_result, data_only=True)
            try:
                self.assertNotIn(
                    "C002",
                    {row["合同号"] for row in _base_rows(workbook["收入分配"]).values()},
                )
            finally:
                workbook.close()

            _append_demand_contract(first_sources[2], "C002", "SC-Z")
            _run(first_sources, third_result, previous=second_result)
            workbook = load_workbook(third_result, data_only=True)
            try:
                rows = _base_rows(workbook["收入分配"])
                actual = rows[("C002", "SC-Z")]
                self.assertIsNone(actual["上期手工金额"])
                self.assertIsNone(actual["手工分配金额"])
                self.assertIsNone(actual["分配备注"])
            finally:
                workbook.close()

    def test_new_result_persists_explicit_row_kind_in_hidden_metadata(
        self,
    ) -> None:
        with TemporaryDirectory() as temporary:
            directory = Path(temporary)
            sources = _write_sources(directory, "explicit", variant="first")
            output = directory / "result.xlsx"
            _run(sources, output)

            issues = IssueLog()
            previous = PreviousResultReader().read(
                output, load_config(CONFIG), issues
            )
            row = next(
                item
                for item in previous.fulfillment_projections
                if item.contract_no == "C002"
            )
            self.assertEqual(CONTRACT_ONLY_NO_DEMAND, row.row_kind)
            self.assertNotIn(
                "PREVIOUS_ROW_KIND_UNAVAILABLE",
                {issue.code for issue in issues.items},
            )

    def test_schema_v2_result_infers_no_demand_from_visible_fields(
        self,
    ) -> None:
        with TemporaryDirectory() as temporary:
            directory = Path(temporary)
            sources = _write_sources(directory, "legacy", variant="first")
            output = directory / "legacy-v08.xlsx"
            _write_v08_no_demand_result(output)

            previous = ExcelInputAdapter().read_previous(
                output, load_config(CONFIG), IssueLog()
            )

            self.assertEqual(
                CONTRACT_ONLY_NO_DEMAND,
                previous.rows[business_key_identity("C002", None)].row_kind,
            )

    def test_blank_month_demand_to_no_demand_is_written_to_both_change_sheets(
        self,
    ) -> None:
        with TemporaryDirectory() as temporary:
            directory = Path(temporary)
            first_sources = _write_sources(directory, "first", variant="first")
            second_sources = _write_sources(directory, "second", variant="second")
            previous = directory / "previous.xlsx"
            output = directory / "current.xlsx"
            _run(first_sources, previous)
            _clear_result_months(previous, "C003", "SC-C")

            result = _run(second_sources, output, previous=previous)

            self.assertGreaterEqual(result.rpd_change_count, 1)
            self.assertGreaterEqual(result.cpd_change_count, 1)
            workbook = load_workbook(output, data_only=True)
            try:
                for mode, sheet_name in (
                    ("RPD", "RPD跨月变化"),
                    ("CPD", "CPD跨月变化"),
                ):
                    row = _base_rows(workbook[sheet_name])[("C003", "SC-C")]
                    self.assertEqual("变为不要货", row["变化方向"])
                    self.assertIsNone(row[f"上期收入年月（按{mode}）"])
                    self.assertIsNone(row[f"本期收入年月（按{mode}）"])
            finally:
                workbook.close()


def _row(role: str, overrides: dict[str, object]) -> ParsedRow:
    config = load_config(CONFIG)
    values = {field: None for field in config.fields[role]}
    values.update(overrides)
    return ParsedRow(
        role=role,
        workbook=f"{role}.xlsx",
        sheet="Sheet1",
        row_number=3,
        values=values,
        raw_values=dict(values),
    )


def _append_monthly_contract(path: Path, contract: str, amount: int) -> None:
    workbook = load_workbook(path)
    try:
        sheet = workbook["当月订货"]
        sheet.append(
            [
                contract,
                "地区M",
                "中国",
                "项目M",
                "BG-M",
                amount,
            ]
        )
        workbook.save(path)
    finally:
        workbook.close()


def _clear_all_supply_centers(path: Path, contract: str) -> None:
    workbook = load_workbook(path)
    try:
        sheet = workbook["要货明细"]
        headers = {cell.value: cell.column for cell in sheet[2]}
        for row_number in range(3, sheet.max_row + 1):
            if sheet.cell(row_number, headers["原合同号"]).value == contract:
                sheet.cell(row_number, headers["供应中心简称"]).value = None
        workbook.save(path)
    finally:
        workbook.close()


def _append_demand_contract(path: Path, contract: str, center: str) -> None:
    workbook = load_workbook(path)
    try:
        sheet = workbook["要货明细"]
        sheet.append(
            [
                contract,
                "地区2D",
                "中国",
                "客户2D",
                "项目2D",
                "有效",
                center,
                "EXW",
                "Y",
                "Y",
                None,
                None,
                date(2026, 8, 1),
                date(2026, 8, 2),
                "BG-2D",
            ]
        )
        workbook.save(path)
    finally:
        workbook.close()


def _clear_result_months(path: Path, contract: str, center: str) -> None:
    workbook = load_workbook(path)
    try:
        sheet = workbook["_fulfillment_projection"]
        headers = {cell.value: cell.column for cell in sheet[1]}
        for row_number in range(2, sheet.max_row + 1):
            if (
                sheet.cell(row_number, headers["合同号"]).value == contract
                and sheet.cell(row_number, headers["履行供应中心"]).value
                == center
            ):
                sheet.cell(
                    row_number, headers["收入年月（按RPD）"]
                ).value = None
                sheet.cell(
                    row_number, headers["收入年月（按CPD）"]
                ).value = None
                break
        workbook.save(path)
    finally:
        workbook.close()


def _contract_rows(sheet) -> dict[str, dict[str, object]]:
    headers = [cell.value for cell in sheet[1]]
    return {
        row["合同号"]: row
        for values in sheet.iter_rows(min_row=2, values_only=True)
        if (row := dict(zip(headers, values))).get("合同号")
    }


def _write_v08_no_demand_result(path: Path) -> None:
    from openpyxl import Workbook
    from tests.test_pipeline import EXPECTED_BASE_HEADERS

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "基表"
    sheet.append(EXPECTED_BASE_HEADERS)
    values = {name: None for name in EXPECTED_BASE_HEADERS}
    values.update(
        {
            "合同号": "C002",
            "遗留量": 10,
            "当月新订货": 0,
            "履行供应中心": None,
            "多个供应中心发货": "N",
            "分批发货": "N",
            "多次要货": "N",
            "分批供应": "N",
            "收入分段类别": "不要货",
        }
    )
    sheet.append([values[name] for name in EXPECTED_BASE_HEADERS])
    metadata = workbook.create_sheet("_tool_meta")
    metadata.append(["schema_version", "2"])
    metadata.append(["base_sheet", "基表"])
    workbook.save(path)
    workbook.close()


if __name__ == "__main__":
    unittest.main()
