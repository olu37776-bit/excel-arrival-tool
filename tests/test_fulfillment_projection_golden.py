from datetime import date
from decimal import Decimal
from pathlib import Path
from tempfile import TemporaryDirectory
import unittest

from openpyxl import load_workbook

from revenue_tool.adapters.excel_reader import ExcelInputAdapter
from revenue_tool.application.pipeline import build_phase1_models
from revenue_tool.config import load_config
from revenue_tool.domain.models import IssueLog, ParsedRow, PreviousData, SourceData, SourceFiles
from revenue_tool.services.calculation import RevenueEngine
from revenue_tool.services.legacy_projection_adapter import (
    LegacyProjectionAdapter,
    compare_golden_rows,
)
from tests.test_pipeline import (
    CONFIG as CONFIG_PATH,
    _set_matching_cell,
    _set_transit_value,
    _write_sources,
)


CONFIG = load_config(CONFIG_PATH)


class FulfillmentProjectionGoldenTest(unittest.TestCase):
    def test_standard_fixture_matches_all_automatic_fields(self) -> None:
        with TemporaryDirectory() as temporary:
            sources = _write_sources(
                Path(temporary), "golden", variant="first"
            )
            source = _read_source(sources)

            differences = _compare(source)

            self.assertEqual([], differences)

    def test_rich_edge_fixture_matches_all_automatic_fields(self) -> None:
        with TemporaryDirectory() as temporary:
            sources = _write_sources(
                Path(temporary), "edge", variant="first"
            )
            _set_matching_cell(
                sources[0],
                "遗留量",
                "华为合同号",
                "C003",
                "设备订未收-新",
                "#VALUE!",
            )
            _set_matching_cell(
                sources[2],
                "要货明细",
                "原合同号",
                "C003",
                "天_ATA",
                "VALUE",
            )
            _set_transit_value(sources[3], "日本", "SC-C", "bad")
            _clear_centers(sources[2], "C004")
            _set_demand_status(sources[2], 5, "运输中")
            _append_fca_record(sources[2])
            source = _read_source(sources)

            differences = _compare(source)

            self.assertEqual([], differences)

    def test_optional_monthly_source_matches_all_automatic_fields(self) -> None:
        with TemporaryDirectory() as temporary:
            sources = _write_sources(
                Path(temporary), "optional", variant="first"
            )
            source = ExcelInputAdapter().read_source(
                SourceFiles(
                    legacy=sources[0],
                    monthly_order=None,
                    demand_detail=sources[2],
                    transit=sources[3],
                ),
                CONFIG,
                IssueLog(),
            )

            differences = _compare(source)

            self.assertEqual([], differences)

    def test_issue_20_is_the_only_precise_v08_difference(self) -> None:
        source = _issue_20_source()
        legacy = RevenueEngine(legacy_carryover_compat=True).calculate(
            source,
            PreviousData({}, usable=False),
            CONFIG,
            IssueLog(),
        )
        models = build_phase1_models(source, CONFIG, IssueLog())
        phase1 = LegacyProjectionAdapter().to_base_rows(
            models.contract_facts,
            models.fulfillment_projections,
        )

        differences = compare_golden_rows(legacy, phase1)

        self.assertEqual(1, len(differences))
        difference = differences[0]
        self.assertEqual(("C020", "sc-a"), difference.business_key)
        self.assertEqual("carryover_type", difference.field)
        self.assertIsNone(difference.legacy_value)
        self.assertEqual("交付类", difference.phase1_value)


def _read_source(sources: tuple[Path, Path, Path, Path]) -> SourceData:
    return ExcelInputAdapter().read_source(
        SourceFiles(*sources), CONFIG, IssueLog()
    )


def _compare(source: SourceData):
    legacy = RevenueEngine().calculate(
        source,
        PreviousData({}, usable=False),
        CONFIG,
        IssueLog(),
    )
    models = build_phase1_models(source, CONFIG, IssueLog())
    phase1 = LegacyProjectionAdapter().to_base_rows(
        models.contract_facts,
        models.fulfillment_projections,
    )
    return compare_golden_rows(legacy, phase1)


def _issue_20_source() -> SourceData:
    legacy_values = {
        "contract_no": "C020",
        "country": None,
        "legacy_amount": Decimal("10.00"),
    }
    demand_values = {
        "contract_no": "C020",
        "country": "阿拉伯联合酋长国",
        "supply_center": "SC-A",
        "incoterm": "EXW",
        "rpd": date(2026, 1, 1),
        "cpd": date(2026, 1, 2),
    }
    return SourceData(
        {
            "legacy": Path("legacy.xlsx"),
            "demand_detail": Path("demand.xlsx"),
            "transit": Path("transit.xlsx"),
        },
        {
            "legacy": [_parsed("legacy", 2, legacy_values)],
            "monthly_order": [],
            "demand_detail": [
                _parsed("demand_detail", 2, demand_values)
            ],
            "transit": [],
        },
        {
            "legacy": "Legacy",
            "demand_detail": "Demand",
            "transit": "Transit",
        },
    )


def _parsed(role: str, row_number: int, overrides: dict) -> ParsedRow:
    values = {field: None for field in CONFIG.fields[role]}
    values.update(overrides)
    return ParsedRow(
        role=role,
        workbook=f"{role}.xlsx",
        sheet=role,
        row_number=row_number,
        values=values,
        raw_values=dict(values),
    )


def _clear_centers(path: Path, contract: str) -> None:
    workbook = load_workbook(path)
    try:
        sheet = workbook["要货明细"]
        headers = {cell.value: cell.column for cell in sheet[2]}
        for row_number in range(3, sheet.max_row + 1):
            if sheet.cell(row_number, headers["原合同号"]).value == contract:
                sheet.cell(row_number, headers["供应中心简称"]).value = None
        workbook.save(path)
    finally:
        workbook.close()


def _append_fca_record(path: Path) -> None:
    workbook = load_workbook(path)
    try:
        sheet = workbook["要货明细"]
        sheet.append(
            [
                "C008",
                "地区8",
                "新加坡",
                "客户8",
                "项目8",
                "有效",
                "SC-FCA",
                "FCA",
                "Y",
                "Y",
                None,
                None,
                date(2026, 8, 1),
                date(2026, 8, 2),
                "BG-8",
            ]
        )
        workbook.save(path)
    finally:
        workbook.close()


def _set_demand_status(path: Path, row_number: int, value: str) -> None:
    workbook = load_workbook(path)
    try:
        sheet = workbook["要货明细"]
        headers = {cell.value: cell.column for cell in sheet[2]}
        sheet.cell(row_number, headers["需求状态"]).value = value
        workbook.save(path)
    finally:
        workbook.close()


if __name__ == "__main__":
    unittest.main()
