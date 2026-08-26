from pathlib import Path
from tempfile import TemporaryDirectory
import unittest

from openpyxl import Workbook

from revenue_tool.adapters.excel_reader import ExcelInputAdapter
from revenue_tool.adapters.sheet_locator import resolve_role_sheet
from revenue_tool.config import load_config
from revenue_tool.domain.models import IssueLog, SourceFiles


ROOT = Path(__file__).resolve().parents[1]
CONFIG_PATH = ROOT / "config" / "default.json"


class SheetLocatorTest(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.config = load_config(CONFIG_PATH)

    def test_all_four_roles_find_field_contract_on_sheet2(self) -> None:
        for role in self.config.sheets:
            with self.subTest(role=role):
                workbook = Workbook()
                try:
                    workbook.active.title = "Sheet1"
                    workbook.active.append(["说明页"])
                    data = workbook.create_sheet("Sheet2")
                    data.append(["导出说明"])
                    data.append(self._headers(role))
                    workbook.create_sheet("Sheet3").append(["无关内容"])

                    resolution = resolve_role_sheet(
                        workbook, role, self.config
                    )

                    self.assertEqual("unique", resolution.mode)
                    self.assertIsNotNone(resolution.selected)
                    self.assertEqual("Sheet2", resolution.selected.sheet_name)
                    self.assertEqual(2, resolution.selected.header_row)
                finally:
                    workbook.close()

    def test_canonical_sheet_name_remains_supported(self) -> None:
        workbook = Workbook()
        try:
            workbook.active.title = self.config.sheets["monthly_order"][
                "canonical"
            ]
            workbook.active.append(self._headers("monthly_order"))

            resolution = resolve_role_sheet(
                workbook, "monthly_order", self.config
            )

            self.assertEqual("unique", resolution.mode)
            self.assertEqual("当月订货", resolution.selected.sheet_name)
        finally:
            workbook.close()

    def test_multiple_matching_sheets_are_ambiguous(self) -> None:
        workbook = Workbook()
        try:
            workbook.active.title = "Sheet1"
            workbook.active.append(self._headers("legacy"))
            workbook.create_sheet("Sheet2").append(self._headers("legacy"))

            resolution = resolve_role_sheet(workbook, "legacy", self.config)

            self.assertEqual("ambiguous", resolution.mode)
            self.assertIsNone(resolution.selected)
            self.assertEqual(
                {"Sheet1", "Sheet2"},
                {item.sheet_name for item in resolution.matches},
            )
        finally:
            workbook.close()

    def test_no_matching_sheet_reports_role_sheets_and_missing_fields(self) -> None:
        with TemporaryDirectory() as temporary:
            directory = Path(temporary)
            paths = {
                role: directory / f"{role}.xlsx"
                for role in self.config.sheets
            }
            for role, path in paths.items():
                workbook = Workbook()
                try:
                    workbook.active.title = "Sheet1"
                    if role == "monthly_order":
                        workbook.active.append(["说明页"])
                        workbook.create_sheet("Sheet3").append(["无关内容"])
                    else:
                        workbook.active.append(self._headers(role))
                    workbook.save(path)
                finally:
                    workbook.close()

            issues = IssueLog()
            source = ExcelInputAdapter().read_source(
                SourceFiles(
                    paths["legacy"],
                    paths["monthly_order"],
                    paths["demand_detail"],
                    paths["transit"],
                ),
                self.config,
                issues,
            )

            matching = [
                issue
                for issue in issues.items
                if issue.code == "SHEET_ROLE_NOT_FOUND"
            ]
            self.assertEqual(1, len(matching))
            issue = matching[0]
            self.assertEqual("monthly_order.xlsx", issue.workbook)
            self.assertEqual("monthly_order", issue.field)
            self.assertIn("Sheet1", issue.message)
            self.assertIn("Sheet3", issue.message)
            self.assertIn("contract_no(华为合同号)", issue.message)
            self.assertIn(
                "monthly_new_order(设备订货(不含VAT))", issue.message
            )
            self.assertNotIn("monthly_order", source.sheet_names)

    def test_adapter_emits_ambiguous_sheet_role_without_selecting(self) -> None:
        with TemporaryDirectory() as temporary:
            directory = Path(temporary)
            paths = {
                role: directory / f"{role}.xlsx"
                for role in self.config.sheets
            }
            for role, path in paths.items():
                workbook = Workbook()
                try:
                    workbook.active.title = "Sheet1"
                    workbook.active.append(self._headers(role))
                    if role == "demand_detail":
                        workbook.create_sheet("Sheet2").append(
                            self._headers(role)
                        )
                    workbook.save(path)
                finally:
                    workbook.close()

            issues = IssueLog()
            source = ExcelInputAdapter().read_source(
                SourceFiles(
                    paths["legacy"],
                    paths["monthly_order"],
                    paths["demand_detail"],
                    paths["transit"],
                ),
                self.config,
                issues,
            )

            matching = [
                issue
                for issue in issues.items
                if issue.code == "AMBIGUOUS_SHEET_ROLE"
            ]
            self.assertEqual(1, len(matching))
            self.assertEqual("demand_detail", matching[0].field)
            self.assertIn("Sheet1@header=1", matching[0].raw_value)
            self.assertIn("Sheet2@header=1", matching[0].raw_value)
            self.assertNotIn("demand_detail", source.sheet_names)

    def _headers(self, role: str) -> list[str]:
        return [
            field["canonical"]
            for field in self.config.fields[role].values()
        ]


if __name__ == "__main__":
    unittest.main()
