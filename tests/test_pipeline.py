from datetime import date
import json
from pathlib import Path
from tempfile import TemporaryDirectory
import unittest

from openpyxl import Workbook, load_workbook

from revenue_tool.application.pipeline import run_pipeline
from revenue_tool.domain.models import WorkbookReadError


ROOT = Path(__file__).resolve().parents[1]
CONFIG = ROOT / "config" / "default.json"


class PipelineTest(unittest.TestCase):
    def test_first_run_matches_documented_grain_and_rules(self) -> None:
        with TemporaryDirectory() as temporary:
            directory = Path(temporary)
            sources = _write_sources(directory, "first", variant="first")
            output = directory / "result.xlsx"

            result = _run(sources, output)

            self.assertEqual(6, result.base_count)
            self.assertEqual(1, result.supply_pull_count)
            workbook = load_workbook(output, data_only=True)
            try:
                self.assertEqual(
                    [
                        "基表",
                        "RPD跨月变化",
                        "CPD跨月变化",
                        "供应需要提拉诉求清单粗表",
                        "异常清单",
                        "_tool_meta",
                    ],
                    workbook.sheetnames,
                )
                self.assertEqual("hidden", workbook["_tool_meta"].sheet_state)
                base = workbook["基表"]
                self.assertEqual(EXPECTED_BASE_HEADERS, [c.value for c in base[1]])
                self.assertEqual("A2", base.freeze_panes)
                self.assertTrue(base.tables)

                rows = _base_rows(base)
                sc_a = rows[("C001", "SC-A")]
                self.assertEqual(100, sc_a["遗留量"])
                self.assertEqual(50, sc_a["当月新订货"])
                self.assertEqual("BG-L", sc_a["BG"])
                self.assertEqual("交付类", sc_a["结转类型"])
                self.assertEqual("Y", sc_a["多个供应中心发货"])
                self.assertEqual("Y|N", sc_a["是否解锁备货"])
                self.assertEqual("Y", sc_a["分批发货"])
                self.assertEqual(30, sc_a["海运周期"])
                self.assertEqual(date(2026, 1, 5), _as_date(sc_a["RPD"]))
                self.assertEqual(date(2026, 1, 10), _as_date(sc_a["最晚RPD"]))
                self.assertIsNone(sc_a["最晚ASD"])
                self.assertEqual("Y", sc_a["货未发完"])
                self.assertEqual("Y", sc_a["多次要货"])
                self.assertEqual("Y", sc_a["分批供应"])
                self.assertEqual(
                    date(2026, 2, 9), _as_date(sc_a["到货日期（按RPD）"])
                )
                self.assertEqual(
                    date(2026, 3, 3), _as_date(sc_a["到货日期（按CPD）"])
                )
                self.assertEqual("2026-02", sc_a["收入年月（按RPD）"])
                self.assertEqual("2026-03", sc_a["收入年月（按CPD）"])

                sc_b = rows[("C001", "SC-B")]
                self.assertEqual(5, sc_b["海运周期"])
                self.assertEqual(
                    date(2026, 2, 6), _as_date(sc_b["到货日期（按RPD）"])
                )

                c003 = rows[("C003", "SC-C")]
                self.assertEqual("客户3D", c003["客户群"])
                self.assertEqual("CIF", c003["贸易术语"])
                self.assertEqual("N", c003["货未发完"])
                self.assertEqual(date(2026, 4, 15), _as_date(c003["最晚ASD"]))
                self.assertEqual(date(2026, 3, 1), _as_date(c003["最晚RPD"]))
                self.assertEqual(
                    date(2026, 5, 5), _as_date(c003["到货日期（按RPD）"])
                )
                self.assertEqual(
                    date(2026, 5, 5), _as_date(c003["到货日期（按CPD）"])
                )

                c007_x = rows[("C007", "SC-X")]
                self.assertEqual("Y", c007_x["货未发完"])
                self.assertEqual("地区7A", c007_x["地区部"])
                self.assertEqual("中国", c007_x["国家"])

                c007_y = rows[("C007", "SC-Y")]
                self.assertEqual(0, c007_y["遗留量"])
                self.assertEqual(0, c007_y["当月新订货"])
                self.assertEqual("未录入订货", c007_y["收入分段类别"])

                c004 = rows[("C004", "SC-D")]
                self.assertEqual(70, c004["当月新订货"])
                self.assertIsNone(c004["收入年月（按CPD）"])

                supply = workbook["供应需要提拉诉求清单粗表"]
                self.assertEqual(EXPECTED_SUPPLY_HEADERS, [c.value for c in supply[1]])
                supply_rows = _rows_by_key(supply)
                self.assertEqual({("C001", "SC-A")}, set(supply_rows))
                self.assertEqual(
                    "2026-02", supply_rows[("C001", "SC-A")]["收入年月（按RPD）"]
                )
                self.assertEqual(
                    "2026-03", supply_rows[("C001", "SC-A")]["收入年月（按CPD）"]
                )

                self.assertEqual(1, workbook["RPD跨月变化"].max_row)
                self.assertEqual(1, workbook["CPD跨月变化"].max_row)
                issue_rows = list(
                    workbook["异常清单"].iter_rows(min_row=2, values_only=True)
                )
                codes = {row[0] for row in issue_rows}
                self.assertIn("DUPLICATE_ROW_IGNORED", codes)
                self.assertIn("CONFLICTING_TRANSIT_DAYS", codes)
                self.assertIn("CONFLICTING_COUNTRY_FOR_CONTRACT", codes)
                self.assertIn("CONTRACT_NOT_FOUND_IN_DEMAND_DETAIL", codes)
                self.assertIn("CONTROL_FLAG_MISMATCH", codes)
                self.assertNotIn("CONFLICTING_CONTRACT_VALUE", codes)
                self.assertNotIn("CONFLICTING_GROUP_VALUE", codes)
                self.assertNotIn("ARRIVAL_CPD_UNAVAILABLE", codes)
                self.assertNotIn("SHIPMENT_STATUS_UNAVAILABLE", codes)
                self.assertNotIn("TEXT_PLACEHOLDER_NORMALIZED_TO_BLANK", codes)
                self.assertNotIn("SUSPECT_AMOUNT_FLOAT_RESIDUE", codes)
                self.assertNotIn("MISSING_INCOTERM", codes)
                control_mismatches = [
                    row for row in issue_rows if row[0] == "CONTROL_FLAG_MISMATCH"
                ]
                self.assertEqual(2, len(control_mismatches))
                self.assertEqual([6, 11], [row[4] for row in control_mismatches])
                contract_missing = next(
                    row
                    for row in issue_rows
                    if row[0] == "CONTRACT_NOT_FOUND_IN_DEMAND_DETAIL"
                )
                self.assertEqual("要货明细未找到该合同号", contract_missing[8])
                transit_conflict = next(
                    row for row in issue_rows if row[0] == "CONFLICTING_TRANSIT_DAYS"
                )
                self.assertEqual("first-transit.xlsx", transit_conflict[2])
                self.assertIn(
                    "first-transit.xlsx/国家运输周期!3=30", transit_conflict[7]
                )
                self.assertIn(
                    "first-transit.xlsx/国家运输周期!5=40", transit_conflict[7]
                )
            finally:
                workbook.close()

    def test_previous_manual_fields_and_two_change_sheets(self) -> None:
        with TemporaryDirectory() as temporary:
            directory = Path(temporary)
            first_sources = _write_sources(directory, "first", variant="first")
            second_sources = _write_sources(directory, "second", variant="second")
            first_result = directory / "first-result.xlsx"
            second_result = directory / "second-result.xlsx"
            _run(first_sources, first_result)
            _set_manual_values(first_result, "C001", "SC-A")

            result = _run(second_sources, second_result, previous=first_result)

            self.assertGreaterEqual(result.rpd_change_count, 3)
            self.assertGreaterEqual(result.cpd_change_count, 3)
            workbook = load_workbook(second_result, data_only=True)
            try:
                base_rows = _base_rows(workbook["基表"])
                inherited = base_rows[("C001", "SC-A")]
                self.assertEqual("Y", inherited["是否手工调整收入月份"])
                self.assertEqual("2026-04", inherited["手工调整收入月份"])
                self.assertEqual("业务确认", inherited["调整备注"])
                self.assertIsNone(
                    base_rows[("C005", "SC-E")]["是否手工调整收入月份"]
                )

                rpd = _rows_by_key(workbook["RPD跨月变化"])
                self.assertEqual("延后", rpd[("C001", "SC-A")]["变化方向"])
                self.assertEqual(1, rpd[("C001", "SC-A")]["变化月数"])
                self.assertEqual("取消", rpd[("C003", "SC-C")]["变化方向"])
                self.assertEqual("新增", rpd[("C005", "SC-E")]["变化方向"])

                cpd = _rows_by_key(workbook["CPD跨月变化"])
                self.assertEqual("提前", cpd[("C001", "SC-A")]["变化方向"])
                self.assertEqual(1, cpd[("C001", "SC-A")]["变化月数"])
                self.assertEqual("取消", cpd[("C003", "SC-C")]["变化方向"])
                self.assertEqual("新增", cpd[("C005", "SC-E")]["变化方向"])
            finally:
                workbook.close()

    def test_monthly_file_is_required_but_contract_match_may_be_absent(self) -> None:
        with TemporaryDirectory() as temporary:
            directory = Path(temporary)
            sources = _write_sources(
                directory,
                "no-match",
                variant="first",
                include_c001_monthly=False,
            )
            output = directory / "result.xlsx"

            _run(sources, output)

            workbook = load_workbook(output, data_only=True)
            try:
                rows = _base_rows(workbook["基表"])
                self.assertEqual(0, rows[("C001", "SC-A")]["当月新订货"])
                codes = {
                    row[0].value
                    for row in workbook["异常清单"].iter_rows(min_row=2)
                }
                self.assertNotIn("MISSING_SHEET", codes)
            finally:
                workbook.close()

    def test_all_source_files_support_auto_named_business_sheets(self) -> None:
        with TemporaryDirectory() as temporary:
            directory = Path(temporary)
            sources = _write_sources(directory, "auto", variant="first")
            for path in sources:
                _move_business_sheet_to_sheet2(path)
            output = directory / "result.xlsx"

            result = _run(sources, output)

            self.assertEqual(6, result.base_count)
            workbook = load_workbook(output, data_only=True)
            try:
                rows = _base_rows(workbook["基表"])
                self.assertEqual(100, rows[("C001", "SC-A")]["遗留量"])
                self.assertEqual(50, rows[("C001", "SC-A")]["当月新订货"])
                codes = {
                    row[0].value
                    for row in workbook["异常清单"].iter_rows(min_row=2)
                }
                self.assertNotIn("SHEET_ROLE_NOT_FOUND", codes)
                self.assertNotIn("AMBIGUOUS_SHEET_ROLE", codes)
            finally:
                workbook.close()

    def test_value_placeholders_are_empty_without_issues(self) -> None:
        with TemporaryDirectory() as temporary:
            directory = Path(temporary)
            sources = _write_sources(directory, "value", variant="first")
            _set_matching_cell(
                sources[0],
                "遗留量",
                "华为合同号",
                "C003",
                "设备订未收-新",
                "#VALUE!",
            )
            _set_matching_cell(
                sources[2],
                "要货明细",
                "原合同号",
                "C003",
                "天_ATA",
                "#VALUE!",
            )
            output = directory / "result.xlsx"

            _run(sources, output)

            workbook = load_workbook(output, data_only=True)
            try:
                c003 = _base_rows(workbook["基表"])[("C003", "SC-C")]
                self.assertEqual(0, c003["遗留量"])
                codes = {
                    row[0].value
                    for row in workbook["异常清单"].iter_rows(min_row=2)
                }
                self.assertNotIn("INVALID_AMOUNT", codes)
                self.assertNotIn("INVALID_DATE", codes)
            finally:
                workbook.close()

    def test_invalid_amount_reports_issue_and_degrades_to_zero(self) -> None:
        with TemporaryDirectory() as temporary:
            directory = Path(temporary)
            sources = _write_sources(directory, "invalid", variant="first")
            _set_matching_cell(
                sources[0],
                "遗留量",
                "华为合同号",
                "C003",
                "设备订未收-新",
                "not-an-amount",
            )
            output = directory / "result.xlsx"

            _run(sources, output)

            workbook = load_workbook(output, data_only=True)
            try:
                c003 = _base_rows(workbook["基表"])[("C003", "SC-C")]
                self.assertEqual(0, c003["遗留量"])
                codes = [
                    row[0].value
                    for row in workbook["异常清单"].iter_rows(min_row=2)
                ]
                self.assertIn("INVALID_AMOUNT", codes)
            finally:
                workbook.close()

    def test_group_text_uses_first_value_without_conflict_issue(self) -> None:
        with TemporaryDirectory() as temporary:
            directory = Path(temporary)
            sources = _write_sources(directory, "text", variant="first")
            _set_matching_cell(
                sources[2],
                "要货明细",
                "原合同号",
                "C003",
                "贸易术语",
                "DAP",
            )
            output = directory / "result.xlsx"

            _run(sources, output)

            workbook = load_workbook(output, data_only=True)
            try:
                c003 = _base_rows(workbook["基表"])[("C003", "SC-C")]
                self.assertEqual("DAP", c003["贸易术语"])
                codes = {
                    row[0].value
                    for row in workbook["异常清单"].iter_rows(min_row=2)
                }
                self.assertNotIn("CONFLICTING_GROUP_VALUE", codes)
            finally:
                workbook.close()

    def test_previous_metadata_survives_display_name_changes(self) -> None:
        with TemporaryDirectory() as temporary:
            directory = Path(temporary)
            first_sources = _write_sources(directory, "first", variant="first")
            second_sources = _write_sources(directory, "second", variant="second")
            first_result = directory / "first-result.xlsx"
            second_result = directory / "second-result.xlsx"
            changed_config = directory / "changed-config.json"
            _run(first_sources, first_result)
            _set_manual_values(first_result, "C001", "SC-A")

            data = json.loads(CONFIG.read_text(encoding="utf-8"))
            data["output"]["sheets"]["base"] = "主数据"
            for column in data["output"]["base_columns"]:
                column["name"] = f"新-{column['id']}"
            changed_config.write_text(
                json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8"
            )

            _run(
                second_sources,
                second_result,
                previous=first_result,
                config=changed_config,
            )

            workbook = load_workbook(second_result, data_only=True)
            try:
                sheet = workbook["主数据"]
                headers = {cell.value: cell.column for cell in sheet[1]}
                row_number = next(
                    row
                    for row in range(2, sheet.max_row + 1)
                    if sheet.cell(row, headers["新-contract_no"]).value == "C001"
                    and sheet.cell(row, headers["新-supply_center"]).value == "SC-A"
                )
                self.assertEqual(
                    "Y", sheet.cell(row_number, headers["新-manual_adjust_flag"]).value
                )
                self.assertEqual(
                    "2026-04",
                    sheet.cell(row_number, headers["新-manual_revenue_month"]).value,
                )
                self.assertEqual(
                    "业务确认", sheet.cell(row_number, headers["新-adjustment_note"]).value
                )
            finally:
                workbook.close()

    def test_previous_29_column_result_remains_usable(self) -> None:
        with TemporaryDirectory() as temporary:
            directory = Path(temporary)
            first_sources = _write_sources(directory, "first", variant="first")
            second_sources = _write_sources(directory, "second", variant="second")
            previous = directory / "previous-29-columns.xlsx"
            output = directory / "result.xlsx"
            _run(first_sources, previous)
            _set_manual_values(previous, "C001", "SC-A")
            _remove_new_base_columns(previous)

            _run(second_sources, output, previous=previous)

            workbook = load_workbook(output, data_only=True)
            try:
                inherited = _base_rows(workbook["基表"])[("C001", "SC-A")]
                self.assertEqual("Y", inherited["是否手工调整收入月份"])
                self.assertEqual("2026-04", inherited["手工调整收入月份"])
            finally:
                workbook.close()

    def test_unusable_previous_does_not_create_false_new_changes(self) -> None:
        with TemporaryDirectory() as temporary:
            directory = Path(temporary)
            sources = _write_sources(directory, "first", variant="first")
            previous = directory / "wrong-previous.xlsx"
            output = directory / "result.xlsx"
            wrong = Workbook()
            wrong.active.title = "错误页"
            wrong.active.append(["错误文件"])
            wrong.save(previous)
            wrong.close()

            _run(sources, output, previous=previous)

            workbook = load_workbook(output, data_only=True)
            try:
                self.assertEqual(1, workbook["RPD跨月变化"].max_row)
                self.assertEqual(1, workbook["CPD跨月变化"].max_row)
                codes = {
                    row[0].value
                    for row in workbook["异常清单"].iter_rows(min_row=2)
                }
                self.assertIn("PREVIOUS_BASE_SHEET_UNAVAILABLE", codes)
            finally:
                workbook.close()

    def test_supply_center_representation_drift_keeps_stable_key(self) -> None:
        with TemporaryDirectory() as temporary:
            directory = Path(temporary)
            first_sources = _write_sources(directory, "first", variant="first")
            second_sources = _write_sources(directory, "second", variant="first")
            _replace_supply_center(second_sources[2], "C001", "SC-A", "　ｓｃ－ａ　")
            first_result = directory / "first-result.xlsx"
            second_result = directory / "second-result.xlsx"
            _run(first_sources, first_result)
            _set_manual_values(first_result, "C001", "SC-A")

            _run(second_sources, second_result, previous=first_result)

            workbook = load_workbook(second_result, data_only=True)
            try:
                inherited = _base_rows(workbook["基表"])[("C001", "sc-a")]
                self.assertEqual("Y", inherited["是否手工调整收入月份"])
                self.assertEqual("2026-04", inherited["手工调整收入月份"])
                self.assertEqual(1, workbook["RPD跨月变化"].max_row)
                self.assertEqual(1, workbook["CPD跨月变化"].max_row)
            finally:
                workbook.close()

    def test_path_guards_cover_four_independent_sources(self) -> None:
        with TemporaryDirectory() as temporary:
            directory = Path(temporary)
            sources = _write_sources(directory, "first", variant="first")
            output = directory / "result.xlsx"

            with self.assertRaisesRegex(ValueError, "互相独立"):
                run_pipeline(
                    sources[0], sources[0], sources[2], sources[3], output, CONFIG
                )
            with self.assertRaisesRegex(ValueError, "任何一个本次源文件"):
                _run(sources, sources[0])
            with self.assertRaisesRegex(ValueError, "上一次成功运行结果"):
                _run(sources, output, previous=output)

            self.assertFalse(output.exists())

    def test_missing_required_source_file_blocks_run(self) -> None:
        with TemporaryDirectory() as temporary:
            directory = Path(temporary)
            sources = list(_write_sources(directory, "first", variant="first"))
            sources[1] = directory / "missing-monthly.xlsx"

            with self.assertRaisesRegex(WorkbookReadError, "工作簿不存在"):
                _run(tuple(sources), directory / "result.xlsx")


EXPECTED_BASE_HEADERS = [
    "合同号", "遗留量", "当月新订货", "BG", "地区部", "国家", "结转类型",
    "客户群", "项目名称", "贸易术语", "履行供应中心", "多个供应中心发货",
    "是否解锁备货", "分批发货", "海运周期", "ATA", "ASD", "RPD",
    "多次要货", "最晚ASD", "最晚RPD", "货未发完", "CPD", "分批供应",
    "到货日期（按RPD）", "到货日期（按CPD）", "收入年月（按RPD）",
    "收入年月（按CPD）", "收入分段类别", "是否手工调整收入月份",
    "手工调整收入月份", "调整备注",
]

EXPECTED_SUPPLY_HEADERS = [
    "合同号", "遗留量", "当月新订货", "地区部", "国家", "客户群",
    "履行供应中心", "收入年月（按RPD）", "收入年月（按CPD）",
]


def _run(
    sources: tuple[Path, Path, Path, Path],
    output: Path,
    *,
    previous: Path | None = None,
    config: Path = CONFIG,
):
    return run_pipeline(*sources, output, config, previous_path=previous)


def _write_sources(
    directory: Path,
    prefix: str,
    *,
    variant: str,
    include_c001_monthly: bool = True,
) -> tuple[Path, Path, Path, Path]:
    legacy_path = directory / f"{prefix}-legacy.xlsx"
    monthly_path = directory / f"{prefix}-monthly.xlsx"
    demand_path = directory / f"{prefix}-demand.xlsx"
    transit_path = directory / f"{prefix}-transit.xlsx"

    legacy_book = Workbook()
    legacy = legacy_book.active
    legacy.title = "遗留量"
    _add_header(legacy, [
        "华为合同号", "地区部中文名称", "国家中文名称", "客户群中文名称",
        "交付项目中文名称", "行销产品维BG中文名称", "设备订未收-新",
    ])
    for row in [
        ["C001", "地区L", "阿拉伯联合酋长国", "客户L", "项目L", "BG-L", 100],
        ["C001", "地区L", "阿拉伯联合酋长国", "客户L", "项目L", "BG-L", 100],
        ["C001", "地区L", "巴西", "客户L", "项目L", "BG-L", 999],
        ["C002", "地区2", "中国", "客户2", "项目2", "BG-2", 2e-13],
        ["C003", "地区3", "日本", "（空白）", "项目3", "BG-3", 300],
    ]:
        legacy.append(row)
    legacy_book.save(legacy_path)
    legacy_book.close()

    monthly_book = Workbook()
    monthly = monthly_book.active
    monthly.title = "当月订货"
    _add_header(monthly, [
        "华为合同号", "地区部中文名称", "国家中文名称", "交付项目中文名称",
        "行销产品维BG中文名称", "设备订货（不含VAT）",
    ])
    if include_c001_monthly:
        monthly.append(["C001", "地区M", "中国", "项目M", "BG-M1", 50])
    monthly.append(["C004", "地区M4", "巴西", "项目M4", "BG-M", 70])
    monthly_book.save(monthly_path)
    monthly_book.close()

    demand_book = Workbook()
    demand = demand_book.active
    demand.title = "要货明细"
    _add_header(demand, [
        "原合同号", "地区部", "国家中文名称", "签约客户群", "交付项目中文名称",
        "需求状态", "供应中心简称", "贸易术语", "备货总控标识", "发货总控标识",
        "天_ATA", "ASD日期", "RPD日期", "CPD日期", "BG_CN",
    ])
    if variant == "first":
        sc_a_1 = [
            "C001", "地区D", "阿拉伯联合酋长国", "客户D", "项目D", "有效",
            "SC-A", "CIF", "Y", "Y", None, None,
            date(2026, 1, 10), date(2026, 1, 20), "BG-D",
        ]
        sc_a_2 = [
            "C001", "地区D", "阿拉伯联合酋长国", "客户D", "项目D", "有效",
            "SC-A", "CIF", "N", "N", None, None,
            date(2026, 1, 5), date(2026, 2, 1), "BG-D",
        ]
        demand.append(sc_a_1)
        demand.append(sc_a_1)
        demand.append(sc_a_2)
        demand.append([
            "C001", "地区D", "阿拉伯联合酋长国", "客户D", "项目D", "有效",
            "SC-B", "fob", "Y", "N", date(2026, 3, 10), "（空白）",
            date(2026, 2, 1), date(2026, 2, 10), "BG-D",
        ])
        demand.append([
            "C003", "地区3D", "日本", "客户3D", "项目3D", "有效", "SC-C",
            None, "Y", "Y", None, date(2026, 4, 15), date(2026, 3, 1),
            date(2026, 3, 5), "BG-3D",
        ])
        demand.append([
            "C003", "地区3D", "日本", "客户3D", "项目3D", "有效", "SC-C",
            "CIF", "Y", "Y", None, date(2026, 4, 15), date(2026, 3, 1),
            date(2026, 3, 5), "BG-3D",
        ])
    else:
        demand.append([
            "C001", "地区D", "阿拉伯联合酋长国", "客户D", "项目D", "有效",
            "SC-A", "CIF", "Y", "Y", None, None,
            date(2026, 2, 10), date(2026, 1, 1), "BG-D",
        ])
        demand.append([
            "C001", "地区D", "阿拉伯联合酋长国", "客户D", "项目D", "有效",
            "SC-A", "CIF", "N", "N", None, None,
            date(2026, 2, 5), date(2026, 1, 15), "BG-D",
        ])
        demand.append([
            "C001", "地区D", "阿拉伯联合酋长国", "客户D", "项目D", "有效",
            "SC-B", "FOB", "Y", "Y", date(2026, 3, 10), None,
            date(2026, 2, 1), date(2026, 2, 10), "BG-D",
        ])
        demand.append([
            "C005", "地区5", "新加坡", "客户5", "项目5", "有效", "SC-E", "EXW",
            "Y", "Y", None, None, date(2026, 6, 1), date(2026, 6, 5), "BG-5",
        ])
    demand.append([
        "C004", "地区4", "巴西", "客户4", "项目4", "有效", "SC-D", "EXW",
        "Y", "Y", None, None, date(2026, 5, 1), None, "BG-4",
    ])
    demand.append([
        "C007", "地区7A", "中国", "客户7A", "项目7A", "有效", "SC-X", "EXW",
        "Y", "Y", None, date(2026, 6, 30), date(2026, 7, 1),
        date(2026, 7, 2), "BG-7A",
    ])
    demand.append([
        "C007", "地区7B", "巴西", "客户7B", "项目7B", "有效", "SC-Y", "EXW",
        "Y", "N", None, None, date(2026, 7, 1), date(2026, 7, 2), "BG-7B",
    ])
    demand_book.save(demand_path)
    demand_book.close()

    transit_book = Workbook()
    transit = transit_book.active
    transit.title = "国家运输周期"
    _add_header(transit, ["国家", "供应中心", "运输周期"])
    transit.append(["阿拉伯联合酋长国", "SC-A", 29.6])
    transit.cell(3, 3).number_format = "#,##0"
    transit.append(["阿拉伯联合酋长国", "SC-A", 29.6])
    transit.cell(4, 3).number_format = "#,##0"
    transit.append(["阿拉伯联合酋长国", "SC-A", 40])
    transit.append(["日本", "SC-C", 20])
    transit_book.save(transit_path)
    transit_book.close()

    return legacy_path, monthly_path, demand_path, transit_path


def _add_header(sheet, headers: list[str]) -> None:
    sheet.append(["业务数据"])
    sheet.append(headers)


def _base_rows(sheet) -> dict[tuple[str, str], dict[str, object]]:
    return _rows_by_key(sheet)


def _rows_by_key(sheet) -> dict[tuple[str, str], dict[str, object]]:
    headers = [cell.value for cell in sheet[1]]
    rows = {}
    for values in sheet.iter_rows(min_row=2, values_only=True):
        data = dict(zip(headers, values))
        rows[(data["合同号"], data["履行供应中心"])] = data
    return rows


def _set_manual_values(path: Path, contract: str, center: str) -> None:
    workbook = load_workbook(path)
    try:
        sheet = workbook["基表"]
        headers = {cell.value: cell.column for cell in sheet[1]}
        for row_number in range(2, sheet.max_row + 1):
            if (
                sheet.cell(row_number, headers["合同号"]).value == contract
                and sheet.cell(row_number, headers["履行供应中心"]).value == center
            ):
                sheet.cell(row_number, headers["是否手工调整收入月份"], "Y")
                sheet.cell(row_number, headers["手工调整收入月份"], "2026-04")
                sheet.cell(row_number, headers["调整备注"], "业务确认")
                break
        workbook.save(path)
    finally:
        workbook.close()


def _replace_supply_center(
    path: Path, contract: str, old_center: str, new_center: str
) -> None:
    workbook = load_workbook(path)
    try:
        sheet = workbook["要货明细"]
        headers = {cell.value: cell.column for cell in sheet[2]}
        for row_number in range(3, sheet.max_row + 1):
            if (
                sheet.cell(row_number, headers["原合同号"]).value == contract
                and sheet.cell(row_number, headers["供应中心简称"]).value == old_center
            ):
                sheet.cell(row_number, headers["供应中心简称"], new_center)
        workbook.save(path)
    finally:
        workbook.close()


def _set_matching_cell(
    path: Path,
    sheet_name: str,
    contract_header: str,
    contract: str,
    field_header: str,
    value,
) -> None:
    workbook = load_workbook(path)
    try:
        sheet = workbook[sheet_name]
        headers = {cell.value: cell.column for cell in sheet[2]}
        for row_number in range(3, sheet.max_row + 1):
            if sheet.cell(row_number, headers[contract_header]).value == contract:
                sheet.cell(row_number, headers[field_header], value)
                break
        workbook.save(path)
    finally:
        workbook.close()


def _move_business_sheet_to_sheet2(path: Path) -> None:
    workbook = load_workbook(path)
    try:
        business = workbook.active
        business.title = "Sheet2"
        intro = workbook.create_sheet("Sheet1", 0)
        intro.append(["文件说明"])
        workbook.create_sheet("Sheet3").append(["无关内容"])
        workbook.save(path)
    finally:
        workbook.close()


def _remove_new_base_columns(path: Path) -> None:
    workbook = load_workbook(path)
    try:
        workbook["基表"].delete_cols(20, 3)
        workbook.save(path)
    finally:
        workbook.close()


def _as_date(value):
    return value.date() if hasattr(value, "date") else value


if __name__ == "__main__":
    unittest.main()
