from __future__ import annotations

from pathlib import Path
from tempfile import TemporaryDirectory
import unittest

from openpyxl import Workbook, load_workbook

from revenue_tool.adapters.previous_result_reader import PreviousResultReader
from revenue_tool.config import load_config
from revenue_tool.domain.models import IssueLog
from revenue_tool.domain.revenue_models import (
    MANUAL_AMOUNT_UNAVAILABLE,
    PREVIOUS_SOURCE_V08,
)
from tests.test_pipeline import (
    CONFIG,
    EXPECTED_BASE_HEADERS,
    _run,
    _set_matching_cell,
    _write_sources,
)


class RevenueAllocationE2ETest(unittest.TestCase):
    def test_overallocation_roundtrip_is_pending_and_not_truncated(self) -> None:
        with TemporaryDirectory() as temporary:
            directory = Path(temporary)
            sources = _write_sources(directory, "over", variant="first")
            first = directory / "over-first.xlsx"
            second = directory / "over-second.xlsx"
            _run(sources, first)
            _edit_allocations_with_shifted_columns(
                first,
                {
                    ("C001", "SC-A"): (100, "over A"),
                    ("C001", "SC-B"): (100, "over B"),
                },
            )

            _run(sources, second, previous=first)

            workbook = load_workbook(second, data_only=True)
            try:
                contract = _contract_rows(workbook["合同收入预测"])["C001"]
                self.assertEqual(200, contract["已分配金额"])
                self.assertEqual(-50, contract["待分配金额"])
                self.assertEqual("分配超额", contract["分配状态"])
                self.assertEqual(0, contract["RPD已归月金额"])
                self.assertEqual(200, contract["RPD待归月金额"])
                details = [
                    row
                    for row in _sheet_rows(workbook["收入归月明细"])
                    if row["合同号"] == "C001"
                ]
                self.assertTrue(details)
                self.assertTrue(
                    all(row["正式归月金额"] == 0 for row in details)
                )
                self.assertIn(
                    "ALLOCATION_EXCEEDS_FORECAST",
                    {
                        row["异常代码"]
                        for row in _sheet_rows(workbook["异常清单"])
                    },
                )
            finally:
                workbook.close()

    def test_negative_manual_amount_roundtrips_without_sign_loss(self) -> None:
        with TemporaryDirectory() as temporary:
            directory = Path(temporary)
            sources = _write_sources(directory, "negative", variant="first")
            _set_matching_cell(
                sources[0],
                "遗留量",
                "华为合同号",
                "C003",
                "设备订未收-新",
                -100,
            )
            first = directory / "negative-first.xlsx"
            second = directory / "negative-second.xlsx"
            _run(sources, first)
            _edit_allocations_with_shifted_columns(
                first, {("C003", "SC-C"): (-40, "negative partial")}
            )

            _run(sources, second, previous=first)

            workbook = load_workbook(second, data_only=True)
            try:
                allocation = _keyed_rows(workbook["收入分配"])[
                    ("C003", "SC-C")
                ]
                self.assertEqual(-40, allocation["手工分配金额"])
                self.assertEqual(-40, allocation["最终金额"])
                contract = _contract_rows(workbook["合同收入预测"])["C003"]
                self.assertEqual(-100, contract["收入预测"])
                self.assertEqual(-40, contract["已分配金额"])
                self.assertEqual(-60, contract["待分配金额"])
                self.assertEqual("部分分配", contract["分配状态"])
            finally:
                workbook.close()

    def test_manual_partial_roundtrip_rebuilds_both_monthly_views(self) -> None:
        with TemporaryDirectory() as temporary:
            directory = Path(temporary)
            sources = _write_sources(directory, "roundtrip", variant="first")
            first = directory / "first.xlsx"
            second = directory / "second.xlsx"
            _run(sources, first)
            _edit_allocations_with_shifted_columns(
                first,
                {
                    ("C001", "SC-A"): (0, "explicit zero"),
                    ("C001", "SC-B"): (40, "partial allocation"),
                },
            )

            result = _run(sources, second, previous=first)

            self.assertEqual(410, result.allocated_amount)
            self.assertEqual(110, result.unallocated_amount)
            workbook = load_workbook(second, data_only=True)
            try:
                allocations = _keyed_rows(workbook["收入分配"])
                sc_a = allocations[("C001", "SC-A")]
                sc_b = allocations[("C001", "SC-B")]
                self.assertEqual(0, sc_a["上期手工金额"])
                self.assertEqual(0, sc_a["手工分配金额"])
                self.assertEqual("explicit zero", sc_a["分配备注"])
                self.assertEqual(40, sc_b["上期手工金额"])
                self.assertEqual(40, sc_b["手工分配金额"])
                self.assertEqual(40, sc_b["最终金额"])
                self.assertEqual("部分分配", sc_b["分配状态"])
                self.assertTrue(sc_b["继承来源run_id"])

                contracts = _contract_rows(workbook["合同收入预测"])
                c001 = contracts["C001"]
                self.assertEqual(40, c001["已分配金额"])
                self.assertEqual(110, c001["待分配金额"])
                self.assertEqual("部分分配", c001["分配状态"])
                for row in contracts.values():
                    self.assertEqual(
                        row["收入预测"],
                        row["已分配金额"] + row["待分配金额"],
                    )
                    self.assertEqual(
                        row["收入预测"],
                        row["RPD已归月金额"]
                        + row["RPD待归月金额"]
                        + row["待分配金额"],
                    )
                    self.assertEqual(
                        row["收入预测"],
                        row["CPD已归月金额"]
                        + row["CPD待归月金额"]
                        + row["待分配金额"],
                    )
                total_forecast = sum(
                    row["收入预测"] for row in contracts.values()
                )
                self.assertEqual(
                    total_forecast,
                    sum(row["RPD已归月金额"] for row in contracts.values())
                    + sum(row["RPD待归月金额"] for row in contracts.values())
                    + sum(row["待分配金额"] for row in contracts.values()),
                )
                self.assertEqual(
                    total_forecast,
                    sum(row["CPD已归月金额"] for row in contracts.values())
                    + sum(row["CPD待归月金额"] for row in contracts.values())
                    + sum(row["待分配金额"] for row in contracts.values()),
                )

                details = _sheet_rows(workbook["收入归月明细"])
                sc_b_details = [
                    row
                    for row in details
                    if row["candidate ID"] == sc_b["candidate ID"]
                ]
                self.assertEqual({"RPD", "CPD"}, {
                    row["统计口径"] for row in sc_b_details
                })
                self.assertTrue(
                    all(row["最终分配金额"] == 40 for row in sc_b_details)
                )
                self.assertTrue(
                    all(row["正式归月金额"] == 40 for row in sc_b_details)
                )

                rpd = _sheet_rows(workbook["RPD月度收入汇总"])
                cpd = _sheet_rows(workbook["CPD月度收入汇总"])
                for row in rpd + cpd:
                    self.assertEqual(
                        row["当月预测"], row["订未发"] + row["发未收"]
                    )
                self.assertTrue(
                    any(
                        row["收入年月"] == "2026-02"
                        and row["BG"] == "BG-L"
                        and row["当月预测"] == 40
                        for row in rpd
                    )
                )
                self.assertTrue(
                    any(
                        row["收入年月"] == "2026-02"
                        and row["BG"] == "BG-L"
                        and row["当月预测"] == 40
                        for row in cpd
                    )
                )
                pending = _sheet_rows(workbook["待处理收入"])
                self.assertTrue(
                    any(
                        row["合同号"] == "C001"
                        and row["待处理金额"] == 110
                        for row in pending
                    )
                )
            finally:
                workbook.close()

    def test_projection_forecast_changes_and_orphan_are_visible(self) -> None:
        with TemporaryDirectory() as temporary:
            directory = Path(temporary)
            first_sources = _write_sources(directory, "first", variant="first")
            changed_sources = _write_sources(directory, "changed", variant="second")
            first = directory / "first.xlsx"
            changed = directory / "changed.xlsx"
            _run(first_sources, first)
            _edit_allocations_with_shifted_columns(
                first,
                {
                    ("C001", "SC-A"): (40, "retain and review"),
                    ("C003", "SC-C"): (25, "orphan me"),
                },
            )
            _set_matching_cell(
                changed_sources[1],
                "当月订货",
                "华为合同号",
                "C001",
                "设备订货（不含VAT）",
                60,
            )

            _run(changed_sources, changed, previous=first)

            workbook = load_workbook(changed, data_only=True)
            try:
                allocations = _keyed_rows(workbook["收入分配"])
                sc_a = allocations[("C001", "SC-A")]
                self.assertEqual(40, sc_a["手工分配金额"])
                self.assertEqual("Y", sc_a["履行投影变化"])
                self.assertEqual("Y", sc_a["合同收入预测变化"])
                self.assertEqual("Y", sc_a["需复核"])
                self.assertIn("PROJECTION_CHANGED", sc_a["异常摘要"])
                self.assertIn(
                    "CONTRACT_REVENUE_FORECAST_CHANGED", sc_a["异常摘要"]
                )
                self.assertIn(
                    "CANDIDATE_ADDED",
                    allocations[("C005", "SC-E")]["异常摘要"],
                )

                pending = _sheet_rows(workbook["待处理收入"])
                orphan = next(
                    row
                    for row in pending
                    if row["待处理原因"] == "ORPHANED_PREVIOUS_ALLOCATION"
                )
                self.assertEqual("C003", orphan["合同号"])
                self.assertEqual(25, orphan["待处理金额"])
                self.assertEqual(25, orphan["上期手工金额"])
                self.assertEqual("orphan me", orphan["上期分配备注"])
                issues = _sheet_rows(workbook["异常清单"])
                self.assertIn(
                    "ORPHANED_PREVIOUS_ALLOCATION",
                    {row["异常代码"] for row in issues},
                )
            finally:
                workbook.close()

    def test_v08_previous_restores_projection_but_not_allocation(self) -> None:
        with TemporaryDirectory() as temporary:
            path = Path(temporary) / "v08.xlsx"
            _write_v08_result(path)

            state = PreviousResultReader().read(
                path, load_config(CONFIG), IssueLog()
            )

            self.assertEqual(PREVIOUS_SOURCE_V08, state.metadata.source_format)
            self.assertTrue(state.usable_for_projection_comparison)
            self.assertFalse(state.usable_for_allocation_inheritance)
            candidate = next(iter(state.candidates_by_id.values()))
            self.assertEqual(
                MANUAL_AMOUNT_UNAVAILABLE,
                candidate.manual_allocation.amount_state,
            )
            self.assertIsNone(candidate.manual_allocation.amount)


def _edit_allocations_with_shifted_columns(
    path: Path,
    changes: dict[tuple[str, str], tuple[float, str]],
) -> None:
    workbook = load_workbook(path)
    try:
        sheet = workbook["收入分配"]
        sheet.insert_cols(2)
        sheet.cell(1, 2, "非关键辅助列")
        headers = {cell.value: cell.column for cell in sheet[1]}
        for row_number in range(2, sheet.max_row + 1):
            key = (
                sheet.cell(row_number, headers["合同号"]).value,
                sheet.cell(row_number, headers["履行供应中心"]).value,
            )
            if key not in changes:
                continue
            amount, note = changes[key]
            sheet.cell(row_number, headers["手工分配金额"], amount)
            sheet.cell(row_number, headers["分配备注"], note)
        workbook.save(path)
    finally:
        workbook.close()


def _sheet_rows(sheet) -> list[dict[str, object]]:
    headers = [cell.value for cell in sheet[1]]
    return [
        dict(zip(headers, values))
        for values in sheet.iter_rows(min_row=2, values_only=True)
        if any(value is not None for value in values)
    ]


def _keyed_rows(sheet) -> dict[tuple[str, str], dict[str, object]]:
    return {
        (row["合同号"], row["履行供应中心"]): row
        for row in _sheet_rows(sheet)
    }


def _contract_rows(sheet) -> dict[str, dict[str, object]]:
    return {row["合同号"]: row for row in _sheet_rows(sheet)}


def _write_v08_result(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "基表"
    sheet.append(EXPECTED_BASE_HEADERS)
    values = {name: None for name in EXPECTED_BASE_HEADERS}
    values.update(
        {
            "合同号": "C001",
            "遗留量": 100,
            "当月新订货": 50,
            "BG": "BG-L",
            "地区部": "地区L",
            "国家": "阿拉伯联合酋长国",
            "结转类型": "交付类",
            "客户群": "客户L",
            "项目名称": "项目L",
            "贸易术语": "CIF",
            "履行供应中心": "SC-A",
            "多个供应中心发货": "N",
            "是否解锁备货": "未解锁",
            "分批发货": "N",
            "海运周期": 30,
            "RPD": "2026-01-01",
            "多次要货": "N",
            "最晚RPD": "2026-01-01",
            "货未发完": "Y",
            "分批供应": "N",
            "到货日期（按RPD）": "2026-01-31",
            "收入年月（按RPD）": "2026-01",
            "收入分段类别": "订未发",
            "是否手工调整收入月份": "Y",
            "手工调整收入月份": "2026-04",
            "调整备注": "旧备注不能迁移为金额",
        }
    )
    sheet.append([values[name] for name in EXPECTED_BASE_HEADERS])
    meta = workbook.create_sheet("_tool_meta")
    meta.append(["schema_version", "2"])
    meta.append(["base_sheet", "基表"])
    workbook.save(path)
    workbook.close()


if __name__ == "__main__":
    unittest.main()
