from __future__ import annotations

from pathlib import Path
from tempfile import TemporaryDirectory
import unittest
from zipfile import ZipFile

from openpyxl import load_workbook

from revenue_tool.adapters.excel_reader import ExcelInputAdapter
from revenue_tool.config import load_config
from revenue_tool.domain.models import IssueLog, SourceFiles, WorkbookReadError
from tests.test_pipeline import (
    CONFIG,
    _base_rows,
    _run,
    _set_transit_value,
    _write_sources,
)


VISIBLE_SHEETS = [
    "基表",
    "RPD跨月变化",
    "CPD跨月变化",
    "供应需要提拉诉求清单粗表",
    "异常清单",
]


class LatestIntegrationTest(unittest.TestCase):
    def test_required_role_cannot_be_silently_skipped(self) -> None:
        source_files = SourceFiles(
            legacy=None,  # type: ignore[arg-type]
            monthly_order=None,
            demand_detail=Path("demand.xlsx"),
            transit=Path("transit.xlsx"),
        )

        with self.assertRaisesRegex(WorkbookReadError, "必选源文件未提供"):
            ExcelInputAdapter().read_source(
                source_files,
                load_config(CONFIG),
                IssueLog(),
            )

    def test_all_visible_sheets_use_reopenable_worksheet_autofilters(self) -> None:
        with TemporaryDirectory() as temporary:
            directory = Path(temporary)
            sources = _write_sources(directory, "filter", variant="first")
            output = directory / "result.xlsx"
            _run(sources, output)

            for _ in range(2):
                workbook = load_workbook(output)
                try:
                    for sheet_name in VISIBLE_SHEETS:
                        sheet = workbook[sheet_name]
                        self.assertFalse(sheet.tables)
                        self.assertEqual(
                            sheet.calculate_dimension(),
                            sheet.auto_filter.ref,
                        )
                        self.assertEqual("A2", sheet.freeze_panes)

                    base = workbook["基表"]
                    headers = {
                        cell.value: cell.column for cell in base[1]
                    }
                    self.assertEqual(
                        "00D9EAF7", base.cell(2, 1).fill.fgColor.rgb
                    )
                    self.assertEqual(
                        "00FFF2CC",
                        base.cell(
                            2, headers["是否手工调整收入月份"]
                        ).fill.fgColor.rgb,
                    )
                    workbook.save(output)
                finally:
                    workbook.close()

            with ZipFile(output) as archive:
                names = archive.namelist()
                self.assertFalse(
                    any(name.startswith("xl/tables/") for name in names)
                )
                for index in range(1, 6):
                    content = archive.read(
                        f"xl/worksheets/sheet{index}.xml"
                    )
                    self.assertIn(b"<autoFilter", content)

    def test_invalid_transit_is_quiet_when_no_arrival_path_needs_it(self) -> None:
        with TemporaryDirectory() as temporary:
            directory = Path(temporary)
            sources = _write_sources(directory, "no-date", variant="first")
            _clear_contract_dates(sources[2], "C003")
            _set_transit_value(sources[3], "日本", "SC-C", "bad")
            output = directory / "result.xlsx"
            _run(sources, output)

            workbook = load_workbook(output, data_only=True)
            try:
                headers = [
                    cell.value for cell in workbook["异常清单"][1]
                ]
                matching = []
                for values in workbook["异常清单"].iter_rows(
                    min_row=2, values_only=True
                ):
                    issue = dict(zip(headers, values))
                    if (
                        issue["业务键"] == "C003 | SC-C"
                        and issue["异常代码"]
                        in {
                            "INVALID_TRANSIT_DAYS",
                            "TRANSIT_VALUE_UNAVAILABLE",
                            "TRANSIT_NOT_FOUND",
                            "TRANSIT_COUNTRY_MISSING",
                        }
                    ):
                        matching.append(issue)
                self.assertEqual([], matching)
                row = _base_rows(workbook["基表"])[("C003", "SC-C")]
                self.assertIsNone(row["海运周期"])
                self.assertIsNone(row["到货日期（按RPD）"])
                self.assertIsNone(row["到货日期（按CPD）"])
            finally:
                workbook.close()


def _clear_contract_dates(path: Path, contract: str) -> None:
    workbook = load_workbook(path)
    try:
        sheet = workbook["要货明细"]
        headers = {cell.value: cell.column for cell in sheet[2]}
        for row_number in range(3, sheet.max_row + 1):
            if sheet.cell(row_number, headers["原合同号"]).value != contract:
                continue
            for field in ("天_ATA", "ASD日期", "RPD日期", "CPD日期"):
                sheet.cell(row_number, headers[field]).value = None
        workbook.save(path)
    finally:
        workbook.close()


if __name__ == "__main__":
    unittest.main()
