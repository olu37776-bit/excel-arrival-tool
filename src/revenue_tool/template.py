from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from revenue_tool.config import load_config


def create_template(output_path: str | Path, config_dir: str | Path) -> Path:
    config = load_config(config_dir)
    workbook = Workbook()
    workbook.remove(workbook.active)
    for section in ("prd", "shipment", "transit_days"):
        sheet = workbook.create_sheet(config.sheet_names[section])
        headers = list(config.columns(section).values())
        sheet.append(headers)
        for cell in sheet[1]:
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.font = Font(color="FFFFFF", bold=True)
        sheet.freeze_panes = "A2"
        for index, header in enumerate(headers, start=1):
            sheet.column_dimensions[chr(64 + index)].width = max(16, len(header) + 2)
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    workbook.close()
    return path

