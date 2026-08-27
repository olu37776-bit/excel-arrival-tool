from __future__ import annotations

from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from revenue_tool.adapters.excel_reader import ExcelInputAdapter
from revenue_tool.config import ToolConfig
from revenue_tool.domain.models import DEMAND_CENTER, IssueLog
from revenue_tool.domain.revenue_models import (
    CANDIDATE_ID_VERSION,
    MANUAL_AMOUNT_BLANK,
    MANUAL_AMOUNT_UNAVAILABLE,
    MANUAL_AMOUNT_VALUE,
    PREVIOUS_SOURCE_NATIVE,
    PREVIOUS_SOURCE_V08,
    ManualAllocationSnapshot,
    PreviousCandidateState,
    PreviousContractState,
    PreviousRunMetadata,
    PreviousRunState,
)
from revenue_tool.services.candidate_identity import build_candidate_id
from revenue_tool.services.normalization import (
    ZERO_AMOUNT,
    normalize_amount,
    normalize_lookup,
    normalize_text,
)
from revenue_tool.services.previous_run_state import (
    contract_from_v08_row,
    projection_from_record,
    projection_from_v08_row,
)


class PreviousResultReader:
    def read(
        self,
        path: str | Path,
        config: ToolConfig,
        issues: IssueLog,
    ) -> PreviousRunState:
        workbook_path = Path(path)
        workbook = load_workbook(workbook_path, data_only=True)
        try:
            schema = _metadata_schema(workbook)
        finally:
            workbook.close()
        if schema and schema.isdigit() and int(schema) >= 4:
            return self._read_native(workbook_path, issues)
        return self._read_v08(workbook_path, config, issues, schema)

    def _read_native(
        self, workbook_path: Path, issues: IssueLog
    ) -> PreviousRunState:
        workbook = load_workbook(workbook_path, data_only=True)
        try:
            meta = workbook["_tool_meta"]
            values = {
                normalize_text(meta.cell(row, 1).value): meta.cell(row, 2).value
                for row in range(1, min(meta.max_row, 9) + 1)
            }
            datasets, field_names = _read_metadata_mappings(meta)
            required = {
                "contract_forecast",
                "allocation",
                "fulfillment_projection",
            }
            if not required <= set(datasets):
                raise ValueError("native metadata dataset mapping incomplete")
            contract_rows = _read_dataset(
                workbook,
                datasets["contract_forecast"],
                field_names["contract_forecast"],
            )
            allocation_rows = _read_dataset(
                workbook,
                datasets["allocation"],
                field_names["allocation"],
            )
            projection_rows = _read_dataset(
                workbook,
                datasets["fulfillment_projection"],
                field_names["fulfillment_projection"],
            )

            contracts: dict[str, PreviousContractState] = {}
            for row in contract_rows:
                contract_no = normalize_text(row.get("contract_no"))
                if not contract_no:
                    continue
                contracts[contract_no] = PreviousContractState(
                    contract_no=contract_no,
                    legacy_amount=_amount(row.get("legacy_amount")),
                    monthly_new_order=_amount(row.get("monthly_new_order")),
                    revenue_forecast=_amount(row.get("revenue_forecast")),
                    bg=_optional_text(row.get("bg")),
                    region=_optional_text(row.get("region")),
                    country=_optional_text(row.get("country")),
                    carryover_type=_optional_text(row.get("carryover_type")),
                    customer_group=_optional_text(row.get("customer_group")),
                    project_name=_optional_text(row.get("project_name")),
                    demand_state=normalize_text(row.get("demand_state")),
                )

            manual_by_id: dict[str, ManualAllocationSnapshot] = {}
            for row in allocation_rows:
                candidate_id = normalize_text(
                    row.get("allocation_candidate_id")
                )
                if not candidate_id:
                    continue
                raw_amount = row.get("manual_allocated_amount")
                if raw_amount is None or normalize_text(raw_amount) == "":
                    state = MANUAL_AMOUNT_BLANK
                    amount = None
                else:
                    parsed = normalize_amount(raw_amount)
                    if parsed is None:
                        issues.add(
                            "PREVIOUS_MANUAL_AMOUNT_INVALID",
                            "上期手工分配金额无法解析，按空白处理",
                            workbook=workbook_path.name,
                            sheet=datasets["allocation"],
                            business_key=candidate_id,
                            field="manual_allocated_amount",
                            raw_value=raw_amount,
                        )
                        state = MANUAL_AMOUNT_BLANK
                        amount = None
                    else:
                        state = MANUAL_AMOUNT_VALUE
                        amount = parsed
                manual_by_id[candidate_id] = ManualAllocationSnapshot(
                    amount_state=state,
                    amount=amount,
                    note=_optional_text(row.get("allocation_note")),
                    source_run_id=_optional_text(values.get("run_id")),
                )

            projections = []
            candidates: dict[str, PreviousCandidateState] = {}
            for row in projection_rows:
                projection = projection_from_record(row)
                projections.append(projection)
                candidate_id = normalize_text(
                    row.get("allocation_candidate_id")
                )
                if projection.row_kind != DEMAND_CENTER or not candidate_id:
                    continue
                if candidate_id in candidates:
                    issues.add(
                        "PREVIOUS_DUPLICATE_CANDIDATE_ID",
                        "上期结果存在重复candidate ID，已禁用该ID继承",
                        workbook=workbook_path.name,
                        sheet=datasets["fulfillment_projection"],
                        business_key=candidate_id,
                    )
                    candidates.pop(candidate_id, None)
                    continue
                manual = manual_by_id.get(
                    candidate_id,
                    ManualAllocationSnapshot(MANUAL_AMOUNT_BLANK),
                )
                candidates[candidate_id] = PreviousCandidateState(
                    allocation_candidate_id=candidate_id,
                    candidate_id_version=normalize_text(
                        row.get("candidate_id_version")
                    ),
                    contract_no=projection.contract_no,
                    supply_center=str(projection.supply_center or ""),
                    row_kind=projection.row_kind,
                    projection=projection,
                    projection_fingerprint=_optional_text(
                        row.get("projection_fingerprint")
                    ),
                    revenue_month_rpd=projection.revenue_month_rpd,
                    revenue_month_cpd=projection.revenue_month_cpd,
                    revenue_segment=projection.revenue_segment,
                    manual_allocation=manual,
                )
            return PreviousRunState(
                metadata=PreviousRunMetadata(
                    metadata_schema=normalize_text(values.get("schema_version")),
                    candidate_id_version=_optional_text(
                        values.get("candidate_id_version")
                    ),
                    projection_fingerprint_version=_optional_text(
                        values.get("projection_fingerprint_version")
                    ),
                    run_id=_optional_text(values.get("run_id")),
                    source_format=PREVIOUS_SOURCE_NATIVE,
                    rules_version=_optional_text(values.get("rules_version")),
                ),
                fulfillment_projections=tuple(projections),
                candidates_by_id=candidates,
                contracts_by_no=contracts,
                usable_for_projection_comparison=True,
                usable_for_allocation_inheritance=True,
            )
        except Exception as exc:
            issues.add(
                "PREVIOUS_NATIVE_RESULT_INVALID",
                "上期新格式结果无法完整恢复，本期不执行历史继承",
                severity="ERROR",
                workbook=workbook_path.name,
                sheet="_tool_meta",
                raw_value=f"{type(exc).__name__}: {exc}",
            )
            return PreviousRunState.empty()
        finally:
            workbook.close()

    def _read_v08(
        self,
        workbook_path: Path,
        config: ToolConfig,
        issues: IssueLog,
        schema: str | None,
    ) -> PreviousRunState:
        previous = ExcelInputAdapter().read_previous(
            workbook_path, config, issues
        )
        if not previous.usable:
            return PreviousRunState.empty()
        projections = []
        contracts: dict[str, PreviousContractState] = {}
        candidates: dict[str, PreviousCandidateState] = {}
        for row in previous.rows.values():
            projection = projection_from_v08_row(row)
            projections.append(projection)
            contracts.setdefault(
                projection.contract_no, contract_from_v08_row(row)
            )
            if projection.row_kind != DEMAND_CENTER or not projection.supply_center:
                continue
            candidate_id = build_candidate_id(
                projection.contract_no,
                projection.supply_center,
                projection.row_kind,
            )
            candidates[candidate_id] = PreviousCandidateState(
                allocation_candidate_id=candidate_id,
                candidate_id_version=CANDIDATE_ID_VERSION,
                contract_no=projection.contract_no,
                supply_center=projection.supply_center,
                row_kind=projection.row_kind,
                projection=projection,
                projection_fingerprint=None,
                revenue_month_rpd=projection.revenue_month_rpd,
                revenue_month_cpd=projection.revenue_month_cpd,
                revenue_segment=projection.revenue_segment,
                manual_allocation=ManualAllocationSnapshot(
                    MANUAL_AMOUNT_UNAVAILABLE
                ),
            )
        return PreviousRunState(
            metadata=PreviousRunMetadata(
                metadata_schema=schema or "0",
                candidate_id_version=CANDIDATE_ID_VERSION,
                projection_fingerprint_version=None,
                run_id=None,
                source_format=PREVIOUS_SOURCE_V08,
            ),
            fulfillment_projections=tuple(projections),
            candidates_by_id=candidates,
            contracts_by_no=contracts,
            usable_for_projection_comparison=True,
            usable_for_allocation_inheritance=False,
            diagnostic_codes=("PREVIOUS_ALLOCATION_UNAVAILABLE",),
        )


def _metadata_schema(workbook) -> str | None:
    if "_tool_meta" not in workbook.sheetnames:
        return None
    sheet = workbook["_tool_meta"]
    if normalize_text(sheet["A1"].value) != "schema_version":
        return None
    return normalize_text(sheet["B1"].value)


def _read_metadata_mappings(sheet) -> tuple[
    dict[str, str], dict[str, dict[str, str]]
]:
    datasets: dict[str, str] = {}
    fields: dict[str, dict[str, str]] = {}
    mode = None
    for row in range(10, sheet.max_row + 1):
        first = normalize_text(sheet.cell(row, 1).value)
        second = normalize_text(sheet.cell(row, 2).value)
        third = normalize_text(sheet.cell(row, 3).value)
        if first == "dataset_id" and second == "sheet_name":
            mode = "datasets"
            continue
        if (
            first == "field_dataset_id"
            and second == "field_id"
            and third == "display_name"
        ):
            mode = "fields"
            continue
        if not first:
            continue
        if mode == "datasets":
            datasets[first] = second
        elif mode == "fields":
            fields.setdefault(first, {})[second] = third
    return datasets, fields


def _read_dataset(
    workbook, sheet_name: str, names_by_id: dict[str, str]
) -> list[dict[str, Any]]:
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"missing dataset sheet: {sheet_name}")
    sheet = workbook[sheet_name]
    header_values = [normalize_text(cell.value) for cell in sheet[1]]
    indexes: dict[str, int] = {}
    for field_id, display_name in names_by_id.items():
        identity = normalize_lookup(display_name)
        matches = [
            index
            for index, header in enumerate(header_values)
            if normalize_lookup(header) == identity
        ]
        if len(matches) != 1:
            raise ValueError(
                f"field {field_id} not uniquely found in {sheet_name}"
            )
        indexes[field_id] = matches[0]
    rows: list[dict[str, Any]] = []
    for cells in sheet.iter_rows(min_row=2, values_only=True):
        if not any(value is not None for value in cells):
            continue
        rows.append(
            {
                field_id: (
                    cells[index] if index < len(cells) else None
                )
                for field_id, index in indexes.items()
            }
        )
    return rows


def _amount(value: Any) -> Decimal:
    parsed = normalize_amount(value)
    return parsed if parsed is not None else ZERO_AMOUNT


def _optional_text(value: Any) -> str | None:
    text = normalize_text(value)
    return text or None
