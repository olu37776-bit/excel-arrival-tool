from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

from revenue_tool.config import ToolConfig
from revenue_tool.domain.errors import InputValidationError
from revenue_tool.domain.models import (
    PrdRecord,
    PreviousRevenueLine,
    ShipmentRecord,
    TransitRule,
)


@dataclass(frozen=True)
class InputData:
    prd_rows: list[PrdRecord]
    shipment_rows: list[ShipmentRecord]
    transit_rules: list[TransitRule]


class ExcelInputAdapter:
    def read(self, path: str | Path, config: ToolConfig) -> InputData:
        workbook_path = Path(path)
        if not workbook_path.exists():
            raise InputValidationError(f"Input workbook does not exist: {workbook_path}")
        workbook = load_workbook(workbook_path, data_only=True, read_only=True)
        try:
            prd_sheet = _get_sheet(workbook, config.sheet_names["prd"])
            shipment_sheet = _get_sheet(workbook, config.sheet_names["shipment"])
            transit_sheet = _get_sheet(workbook, config.sheet_names["transit_days"])

            prd_rows = _read_prd(prd_sheet, config.columns("prd"))
            shipment_rows = _read_shipments(
                shipment_sheet, config.columns("shipment")
            )
            transit_rules = _read_transit_rules(
                transit_sheet, config.columns("transit_days")
            )
        finally:
            workbook.close()
        return InputData(prd_rows, shipment_rows, transit_rules)

    def read_previous(self, path: str | Path) -> list[PreviousRevenueLine]:
        workbook_path = Path(path)
        if not workbook_path.exists():
            raise InputValidationError(
                f"Previous result workbook does not exist: {workbook_path}"
            )
        workbook = load_workbook(workbook_path, data_only=True, read_only=True)
        try:
            sheet = _get_sheet(workbook, "Revenue Detail")
            header = _header_index(sheet)
            required = {
                "business_key": "Business Key",
                "po_number": "PO Number",
                "contract_number": "Contract Number",
                "shipping_point": "Shipping Point",
                "shipment_id": "Shipment ID",
                "revenue_month": "Revenue Month",
            }
            indexes = _resolve_columns(header, required, set(required))
            result: list[PreviousRevenueLine] = []
            for row_number, values in enumerate(
                sheet.iter_rows(min_row=2, values_only=True), start=2
            ):
                if _is_blank(values):
                    continue
                result.append(
                    PreviousRevenueLine(
                        business_key=_required_text(
                            _cell(values, indexes["business_key"]),
                            sheet.title,
                            row_number,
                            "Business Key",
                        ),
                        po_number=_text(_cell(values, indexes["po_number"])),
                        contract_number=_text(
                            _cell(values, indexes["contract_number"])
                        ),
                        shipping_point=_text(
                            _cell(values, indexes["shipping_point"])
                        ),
                        shipment_id=_text(_cell(values, indexes["shipment_id"])),
                        previous_revenue_month=_normalise_month(
                            _cell(values, indexes["revenue_month"]),
                            sheet.title,
                            row_number,
                            "Revenue Month",
                        ),
                    )
                )
            return result
        finally:
            workbook.close()


def _read_prd(sheet, mapping: dict[str, str]) -> list[PrdRecord]:
    indexes = _resolve_columns(
        _header_index(sheet),
        mapping,
        {"po_number", "prd", "original_po_quantity"},
    )
    result: list[PrdRecord] = []
    errors: list[str] = []
    for row_number, values in enumerate(
        sheet.iter_rows(min_row=2, values_only=True), start=2
    ):
        if _is_blank(values):
            continue
        try:
            result.append(
                PrdRecord(
                    po_number=_required_text(
                        _value(values, indexes, "po_number"),
                        sheet.title,
                        row_number,
                        mapping["po_number"],
                    ),
                    prd=_date_value(
                        _value(values, indexes, "prd"),
                        sheet.title,
                        row_number,
                        mapping["prd"],
                        required=True,
                    ),
                    original_po_quantity=_decimal_value(
                        _value(values, indexes, "original_po_quantity"),
                        sheet.title,
                        row_number,
                        mapping["original_po_quantity"],
                        required=True,
                    ),
                    contract_number=_text(_value(values, indexes, "contract_number")),
                    shipping_point=_text(_value(values, indexes, "shipping_point")),
                )
            )
        except InputValidationError as exc:
            errors.append(str(exc))
    _raise_row_errors(errors)
    if not result:
        raise InputValidationError(f"Sheet '{sheet.title}' contains no PRD data")
    return result


def _read_shipments(sheet, mapping: dict[str, str]) -> list[ShipmentRecord]:
    required = {
        "po_number",
        "plan_date",
        "plan_quantity",
        "contract_number",
        "shipping_point",
        "trade_type",
    }
    indexes = _resolve_columns(_header_index(sheet), mapping, required)
    result: list[ShipmentRecord] = []
    errors: list[str] = []
    for row_number, values in enumerate(
        sheet.iter_rows(min_row=2, values_only=True), start=2
    ):
        if _is_blank(values):
            continue
        try:
            result.append(
                ShipmentRecord(
                    po_number=_required_text(
                        _value(values, indexes, "po_number"),
                        sheet.title,
                        row_number,
                        mapping["po_number"],
                    ),
                    plan_date=_date_value(
                        _value(values, indexes, "plan_date"),
                        sheet.title,
                        row_number,
                        mapping["plan_date"],
                        required=True,
                    ),
                    plan_quantity=_decimal_value(
                        _value(values, indexes, "plan_quantity"),
                        sheet.title,
                        row_number,
                        mapping["plan_quantity"],
                        required=True,
                    ),
                    contract_number=_required_text(
                        _value(values, indexes, "contract_number"),
                        sheet.title,
                        row_number,
                        mapping["contract_number"],
                    ),
                    shipping_point=_required_text(
                        _value(values, indexes, "shipping_point"),
                        sheet.title,
                        row_number,
                        mapping["shipping_point"],
                    ),
                    trade_type=_required_text(
                        _value(values, indexes, "trade_type"),
                        sheet.title,
                        row_number,
                        mapping["trade_type"],
                    ),
                    shipment_id=_text(_value(values, indexes, "shipment_id")),
                    revenue_amount=_decimal_value(
                        _value(values, indexes, "revenue_amount"),
                        sheet.title,
                        row_number,
                        mapping.get("revenue_amount", "Revenue Amount"),
                        required=False,
                    ),
                    source_row=row_number,
                )
            )
        except InputValidationError as exc:
            errors.append(str(exc))
    _raise_row_errors(errors)
    if not result:
        raise InputValidationError(f"Sheet '{sheet.title}' contains no Shipment data")
    return result


def _read_transit_rules(sheet, mapping: dict[str, str]) -> list[TransitRule]:
    required = {"trade_type", "transit_days"}
    indexes = _resolve_columns(_header_index(sheet), mapping, required)
    result: list[TransitRule] = []
    errors: list[str] = []
    for row_number, values in enumerate(
        sheet.iter_rows(min_row=2, values_only=True), start=2
    ):
        if _is_blank(values):
            continue
        try:
            days = _decimal_value(
                _value(values, indexes, "transit_days"),
                sheet.title,
                row_number,
                mapping["transit_days"],
                required=True,
            )
            if days != days.to_integral_value() or days < 0:
                raise InputValidationError(
                    f"{sheet.title}!{mapping['transit_days']} row {row_number}: "
                    "Transit Days must be a non-negative whole number"
                )
            result.append(
                TransitRule(
                    trade_type=_required_text(
                        _value(values, indexes, "trade_type"),
                        sheet.title,
                        row_number,
                        mapping["trade_type"],
                    ),
                    transit_days=int(days),
                )
            )
        except InputValidationError as exc:
            errors.append(str(exc))
    _raise_row_errors(errors)
    if not result:
        raise InputValidationError(f"Sheet '{sheet.title}' contains no transit rules")
    return result


def _get_sheet(workbook, name: str):
    if name not in workbook.sheetnames:
        raise InputValidationError(
            f"Missing sheet '{name}'. Available sheets: {', '.join(workbook.sheetnames)}"
        )
    return workbook[name]


def _header_index(sheet) -> dict[str, int]:
    first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not first_row:
        raise InputValidationError(f"Sheet '{sheet.title}' has no header row")
    result: dict[str, int] = {}
    for index, value in enumerate(first_row):
        if value is not None and str(value).strip():
            result[_normalise_header(value)] = index
    return result


def _resolve_columns(
    header: dict[str, int],
    mapping: dict[str, str],
    required_fields: set[str],
) -> dict[str, int | None]:
    indexes: dict[str, int | None] = {}
    missing: list[str] = []
    for canonical, configured_name in mapping.items():
        index = header.get(_normalise_header(configured_name))
        indexes[canonical] = index
        if canonical in required_fields and index is None:
            missing.append(configured_name)
    if missing:
        raise InputValidationError("Missing required columns: " + ", ".join(missing))
    return indexes


def _value(values: tuple[Any, ...], indexes: dict[str, int | None], field: str):
    index = indexes.get(field)
    return None if index is None else _cell(values, index)


def _cell(values: tuple[Any, ...], index: int | None):
    if index is None or index >= len(values):
        return None
    return values[index]


def _required_text(value: Any, sheet: str, row: int, column: str) -> str:
    result = _text(value)
    if not result:
        raise InputValidationError(f"{sheet}!{column} row {row}: value is required")
    return result


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _decimal_value(
    value: Any,
    sheet: str,
    row: int,
    column: str,
    required: bool,
) -> Decimal | None:
    if value is None or (isinstance(value, str) and not value.strip()):
        if required:
            raise InputValidationError(f"{sheet}!{column} row {row}: value is required")
        return None
    try:
        return Decimal(str(value).replace(",", "").strip())
    except (InvalidOperation, ValueError) as exc:
        raise InputValidationError(
            f"{sheet}!{column} row {row}: invalid number '{value}'"
        ) from exc


def _date_value(
    value: Any,
    sheet: str,
    row: int,
    column: str,
    required: bool,
) -> date | None:
    if value is None or (isinstance(value, str) and not value.strip()):
        if required:
            raise InputValidationError(f"{sheet}!{column} row {row}: value is required")
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            converted = from_excel(value)
            return converted.date() if isinstance(converted, datetime) else converted
        except (TypeError, ValueError, OverflowError):
            pass
    text = str(value).strip()
    for pattern in (
        "%Y-%m-%d",
        "%Y/%m/%d",
        "%Y.%m.%d",
        "%Y-%m",
        "%Y/%m",
        "%b-%y",
        "%b %Y",
    ):
        try:
            return datetime.strptime(text, pattern).date()
        except ValueError:
            continue
    raise InputValidationError(
        f"{sheet}!{column} row {row}: invalid date or month '{value}'"
    )


def _normalise_month(value: Any, sheet: str, row: int, column: str) -> str:
    if isinstance(value, str):
        text = value.strip()
        if len(text) == 7 and text[4] == "-":
            try:
                datetime.strptime(text, "%Y-%m")
                return text
            except ValueError:
                pass
    parsed = _date_value(value, sheet, row, column, required=True)
    return parsed.strftime("%Y-%m")


def _normalise_header(value: Any) -> str:
    return " ".join(str(value).strip().casefold().split())


def _is_blank(values: tuple[Any, ...]) -> bool:
    return all(value is None or str(value).strip() == "" for value in values)


def _raise_row_errors(errors: list[str]) -> None:
    if errors:
        shown = errors[:20]
        suffix = f"\n... and {len(errors) - 20} more" if len(errors) > 20 else ""
        raise InputValidationError("\n".join(shown) + suffix)

