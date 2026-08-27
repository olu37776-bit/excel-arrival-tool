from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from revenue_tool.config import ToolConfig
from revenue_tool.domain.models import (
    BaseRow,
    ComparisonRow,
    IssueLog,
)


_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_EDITABLE_FILL = PatternFill("solid", fgColor="FFF2CC")
_BANDED_FILL = PatternFill("solid", fgColor="D9EAF7")
_THIN_BLUE = Side(style="thin", color="9EADBF")


class ExcelOutputAdapter:
    def write(
        self,
        output_path: str | Path,
        base_rows: list[BaseRow],
        rpd_changes: list[ComparisonRow],
        cpd_changes: list[ComparisonRow],
        supply_pull_rows: list[ComparisonRow],
        issues: IssueLog,
        config: ToolConfig,
    ) -> Path:
        workbook = Workbook()
        workbook.remove(workbook.active)
        sheets = config.output["sheets"]

        base_columns = config.base_columns
        self._write_table_sheet(
            workbook,
            sheets["base"],
            base_columns,
            [row.values for row in base_rows],
            "BaseTable",
            date_fields={
                "ata",
                "asd",
                "rpd",
                "latest_asd",
                "latest_rpd",
                "cpd",
                "arrival_date_rpd",
                "arrival_date_cpd",
            },
            amount_fields={
                "legacy_amount",
                "monthly_new_order",
                "revenue_forecast",
                "manual_revenue_forecast_rpd",
                "manual_revenue_forecast_cpd",
            },
            editable_fields={
                "manual_adjust_flag",
                "manual_revenue_forecast_rpd",
                "manual_revenue_forecast_cpd",
                "manual_revenue_month",
                "adjustment_note",
            },
        )
        self._write_change_sheet(
            workbook,
            sheets["rpd_changes"],
            "RPD",
            rpd_changes,
            config,
            "RPDChangesTable",
        )
        self._write_change_sheet(
            workbook,
            sheets["cpd_changes"],
            "CPD",
            cpd_changes,
            config,
            "CPDChangesTable",
        )
        self._write_table_sheet(
            workbook,
            sheets["supply_pull"],
            config.output["supply_pull_columns"],
            [row.values for row in supply_pull_rows],
            "SupplyPullTable",
            amount_fields={"legacy_amount", "monthly_new_order"},
        )
        self._write_table_sheet(
            workbook,
            sheets["issues"],
            config.output["issue_columns"],
            [issue.as_dict() for issue in issues.items],
            "IssuesTable",
        )
        self._write_metadata_sheet(workbook, config, base_rows)

        path = Path(output_path)
        path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(path)
        workbook.close()
        return path

    def _write_metadata_sheet(
        self,
        workbook: Workbook,
        config: ToolConfig,
        base_rows: list[BaseRow],
    ) -> None:
        sheet = workbook.create_sheet("_tool_meta")
        sheet.append(["schema_version", "3"])
        sheet.append(["base_sheet", config.output["sheets"]["base"]])
        sheet.append([])
        sheet.append(["field_id", "display_name"])
        for column in config.base_columns:
            sheet.append([column["id"], column["name"]])
        sheet.append([])
        sheet.append(
            [
                "row_kind_contract_no",
                "row_kind_supply_center",
                "row_kind",
            ]
        )
        for row in base_rows:
            sheet.append(
                [
                    row.values.get("contract_no"),
                    row.values.get("supply_center"),
                    row.row_kind,
                ]
            )
        sheet.sheet_state = "hidden"

    def _write_change_sheet(
        self,
        workbook: Workbook,
        sheet_name: str,
        mode: str,
        rows: list[ComparisonRow],
        config: ToolConfig,
        table_name: str,
    ) -> None:
        columns = list(config.output["change_common_columns"]) + list(
            config.output["change_tail_columns"][mode.lower()]
        )
        self._write_table_sheet(
            workbook,
            sheet_name,
            columns,
            [row.values for row in rows],
            table_name,
            amount_fields={"legacy_amount", "monthly_new_order"},
        )

    def _write_table_sheet(
        self,
        workbook: Workbook,
        sheet_name: str,
        columns: list[dict[str, str]],
        rows: list[dict[str, Any]],
        _table_name: str,
        *,
        date_fields: set[str] | None = None,
        amount_fields: set[str] | None = None,
        editable_fields: set[str] | None = None,
    ) -> None:
        date_fields = date_fields or set()
        amount_fields = amount_fields or set()
        editable_fields = editable_fields or set()
        sheet = workbook.create_sheet(sheet_name)
        sheet.sheet_view.showGridLines = False
        sheet.freeze_panes = "A2"

        field_ids = [column["id"] for column in columns]
        sheet.append([column["name"] for column in columns])
        for row in rows:
            sheet.append(
                [_excel_value(row.get(field)) for field in field_ids]
            )

        header = sheet[1]
        for cell in header:
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )
            cell.border = Border(bottom=_THIN_BLUE)
        sheet.row_dimensions[1].height = 32

        for index, field in enumerate(field_ids, start=1):
            letter = get_column_letter(index)
            if field in date_fields:
                for cell in sheet[letter][1:]:
                    cell.number_format = "yyyy-mm-dd"
                    cell.alignment = Alignment(horizontal="center")
            elif field in amount_fields:
                for cell in sheet[letter][1:]:
                    cell.number_format = "#,##0.00"
                    cell.alignment = Alignment(horizontal="right")
            else:
                for cell in sheet[letter][1:]:
                    cell.alignment = Alignment(
                        horizontal="left",
                        vertical="top",
                        wrap_text=False,
                    )
            if field in editable_fields and sheet.max_row >= 2:
                for cell in sheet[letter][1:]:
                    cell.fill = _EDITABLE_FILL

        # Structured Table supplied the old row stripes.  Preserve that
        # visual hierarchy after switching to a worksheet AutoFilter, while
        # keeping the three editable columns visibly yellow.
        for row_number in range(2, sheet.max_row + 1):
            if row_number % 2:
                continue
            for column_number, field in enumerate(field_ids, start=1):
                if field not in editable_fields:
                    sheet.cell(row_number, column_number).fill = _BANDED_FILL

        last_column = get_column_letter(len(columns))
        last_row = max(1, sheet.max_row)
        reference = f"A1:{last_column}{last_row}"
        sheet.auto_filter.ref = reference

        for index, column in enumerate(columns, start=1):
            header_length = len(column["name"])
            data_lengths = [
                len(_display_value(row.get(column["id"])))
                for row in rows[:200]
            ]
            width = min(
                38,
                max(12, header_length * 2 + 2, *(data_lengths or [0])),
            )
            sheet.column_dimensions[get_column_letter(index)].width = width


def _excel_value(value: Any) -> Any:
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    if value is None:
        return None
    if isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _display_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (date, datetime)):
        return value.strftime("%Y-%m-%d")
    return str(value)
