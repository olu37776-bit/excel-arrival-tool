from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
import math
from pathlib import Path
import re
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

from revenue_tool.adapters.sheet_locator import resolve_role_sheet
from revenue_tool.config import ToolConfig
from revenue_tool.domain.models import (
    BaseRow,
    CONTRACT_ONLY_NO_DEMAND,
    DEMAND_CENTER,
    IssueLog,
    ParsedRow,
    PreviousData,
    SourceData,
    SourceFiles,
    WorkbookReadError,
)
from revenue_tool.services.field_matching import resolve_name
from revenue_tool.services.normalization import (
    business_key_identity,
    is_business_blank,
    MANUAL_MONTH_INVALID,
    MANUAL_MONTH_YEAR_REQUIRED,
    nonblank,
    normalize_amount,
    normalize_manual_revenue_month,
    normalize_signature_value,
    normalize_text,
    ZERO_AMOUNT,
)


_PREVIOUS_MANUAL_AMOUNT_FIELDS = {
    "manual_revenue_month",
}
_PREVIOUS_MANUAL_MONTH_FIELDS = {
    "manual_revenue_forecast_rpd",
    "manual_revenue_forecast_cpd",
}
_PREVIOUS_MANUAL_FLAG_FIELDS = {"manual_revenue_segment_flag"}

# Display names are not identities.  Metadata-backed results resolve by the
# stable field ID; these aliases keep older workbooks without usable metadata
# readable after Issue #27's display-only rename.
_PREVIOUS_OUTPUT_NAME_ALIASES = {
    "manual_adjust_flag": ("是否手工调整收入月份",),
    "manual_revenue_forecast_rpd": ("手工调整收入预测（按RPD）",),
    "manual_revenue_forecast_cpd": ("手工调整收入预测（按CPD）",),
    "manual_revenue_month": ("手工调整收入月份",),
}


class ExcelInputAdapter:
    def read_source(
        self,
        source_files: SourceFiles,
        config: ToolConfig,
        issues: IssueLog,
    ) -> SourceData:
        paths = source_files.as_dict()
        result: dict[str, list[ParsedRow]] = {role: [] for role in paths}
        sheet_names: dict[str, str] = {}
        for role, workbook_path in paths.items():
            if workbook_path is None:
                if config.sheets[role]["optional"]:
                    continue
                raise WorkbookReadError(f"必选源文件未提供: {role}")
            workbook = _open_workbook(workbook_path)
            try:
                resolution = resolve_role_sheet(workbook, role, config)
                if resolution.mode == "ambiguous":
                    matches = [
                        f"{item.sheet_name}@header={item.header_row}"
                        for item in resolution.matches
                    ]
                    issues.add(
                        "AMBIGUOUS_SHEET_ROLE",
                        f"角色 {role} 同时命中多个业务 Sheet，未自动选择",
                        workbook=workbook_path.name,
                        field=role,
                        raw_value=" | ".join(matches),
                    )
                    continue
                if resolution.mode == "not_found":
                    actual_sheets = " | ".join(workbook.sheetnames)
                    missing_fields = _missing_required_fields_summary(
                        resolution.fingerprints,
                        role,
                        config,
                    )
                    issues.add(
                        "SHEET_ROLE_NOT_FOUND",
                        (
                            "未找到满足字段契约的业务 Sheet；"
                            f"角色={role}；实际 Sheet={actual_sheets}；"
                            f"缺失关键字段={missing_fields}"
                        ),
                        workbook=workbook_path.name,
                        field=role,
                        raw_value=(
                            f"role={role}; sheets={actual_sheets}; "
                            f"missing_required={missing_fields}"
                        ),
                    )
                    continue
                selected = resolution.selected
                if selected is None or selected.header_row is None:
                    raise AssertionError("唯一 Sheet 解析结果必须包含表头行")
                worksheet = workbook[selected.sheet_name]
                sheet_names[role] = worksheet.title
                result[role] = self._read_role_sheet(
                    workbook_path,
                    workbook.epoch,
                    worksheet,
                    role,
                    config,
                    issues,
                    header_row=selected.header_row,
                )
            finally:
                workbook.close()
        return SourceData(
            {
                role: path
                for role, path in paths.items()
                if path is not None
            },
            result,
            sheet_names,
        )

    def read_previous(
        self, path: str | Path, config: ToolConfig, issues: IssueLog
    ) -> PreviousData:
        workbook_path = Path(path)
        workbook = _open_workbook(workbook_path)
        try:
            (
                expected_sheet,
                previous_names,
                previous_row_kinds,
            ) = _read_previous_metadata(
                workbook,
                config,
                workbook_path,
                issues,
            )
            match = resolve_name(
                expected_sheet,
                [],
                list(workbook.sheetnames),
                config.workbook["contains_direction"],
            )
            if match.index is None:
                issues.add(
                    "PREVIOUS_BASE_SHEET_UNAVAILABLE",
                    f"上期结果未唯一找到 Sheet: {expected_sheet}",
                    workbook=workbook_path.name,
                    raw_value=" | ".join(workbook.sheetnames),
                )
                return PreviousData({}, usable=False)
            sheet = workbook[workbook.sheetnames[match.index]]
            header_row = self._detect_output_header(
                sheet,
                config,
                [
                    previous_names.get(column["id"], column["name"])
                    for column in config.base_columns
                ],
            )
            if header_row is None:
                issues.add(
                    "PREVIOUS_HEADER_NOT_FOUND",
                    "上期基表未找到可识别表头",
                    workbook=workbook_path.name,
                    sheet=sheet.title,
                )
                return PreviousData({}, usable=False)
            headers = [
                normalize_text(cell.value)
                for cell in sheet[header_row]
            ]
            indexes: dict[str, int | None] = {}
            for column in config.base_columns:
                previous_name = previous_names.get(
                    column["id"], column["name"]
                )
                match = resolve_name(
                    previous_name,
                    _PREVIOUS_OUTPUT_NAME_ALIASES.get(column["id"], ()),
                    headers,
                    config.workbook["contains_direction"],
                )
                indexes[column["id"]] = match.index
                if match.index is None:
                    issues.add(
                        "PREVIOUS_FIELD_UNAVAILABLE",
                        f"上期基表字段未唯一匹配: {previous_name}",
                        workbook=workbook_path.name,
                        sheet=sheet.title,
                        field=column["id"],
                    )
            essential = {
                "contract_no",
                "supply_center",
                "revenue_month_rpd",
                "revenue_month_cpd",
            }
            if any(indexes[field] is None for field in essential):
                issues.add(
                    "PREVIOUS_BASE_UNUSABLE",
                    "上期基表缺少比较键或自动收入年月字段，本期不执行继承与比较",
                    workbook=workbook_path.name,
                    sheet=sheet.title,
                    field=" | ".join(
                        sorted(
                            field
                            for field in essential
                            if indexes[field] is None
                        )
                    ),
                )
                return PreviousData({}, usable=False)
            rows: dict[tuple[str, str], BaseRow] = {}
            for row_number in range(header_row + 1, sheet.max_row + 1):
                cells = list(sheet[row_number])
                if _row_is_blank(cells):
                    continue
                field_cells = {
                    column["id"]: (
                        cells[index]
                        if index is not None and index < len(cells)
                        else None
                    )
                    for column in config.base_columns
                    for index in (indexes[column["id"]],)
                }
                values: dict[str, Any] = {}
                for field in (
                    "contract_no",
                    "supply_center",
                    "revenue_month_rpd",
                    "revenue_month_cpd",
                ):
                    cell = field_cells[field]
                    values[field] = _parse_previous_value(
                        field,
                        cell.value if cell is not None else None,
                        workbook.epoch,
                    )

                display_key = (
                    str(values.get("contract_no") or ""),
                    str(values.get("supply_center") or ""),
                )
                for field, primary_field, secondary_field in (
                    (
                        "manual_revenue_forecast_rpd",
                        "revenue_month_rpd",
                        "revenue_month_cpd",
                    ),
                    (
                        "manual_revenue_forecast_cpd",
                        "revenue_month_cpd",
                        "revenue_month_rpd",
                    ),
                ):
                    cell = field_cells[field]
                    raw_value = cell.value if cell is not None else None
                    result = normalize_manual_revenue_month(
                        raw_value,
                        primary_reference_month=values.get(primary_field),
                        secondary_reference_month=values.get(secondary_field),
                        data_type=(
                            getattr(cell, "data_type", None)
                            if cell is not None
                            else None
                        ),
                    )
                    values[field] = result.value
                    context = _manual_month_reference_context(
                        values.get(primary_field),
                        values.get(secondary_field),
                    )
                    if result.status == MANUAL_MONTH_YEAR_REQUIRED:
                        issues.add(
                            "MANUAL_MONTH_YEAR_REQUIRED",
                            (
                                "调整月份只填写了月份，但无法根据本行自动"
                                "收入年月唯一确定年份；请填写完整年月，"
                                f"例如2026-09。{context}"
                            ),
                            workbook=workbook_path.name,
                            sheet=sheet.title,
                            row_number=row_number,
                            business_key=_display_key(display_key),
                            field=field,
                            raw_value=raw_value,
                        )
                    elif result.status == MANUAL_MONTH_INVALID:
                        issues.add(
                            "INVALID_PREVIOUS_MANUAL_MONTH",
                            (
                                "调整月份无法识别。可填写2026-09、"
                                "2026年9月、9月或9；仅填写月份时，"
                                "系统会根据本行自动收入年月补全年份。"
                                f"{context}"
                            ),
                            workbook=workbook_path.name,
                            sheet=sheet.title,
                            row_number=row_number,
                            business_key=_display_key(display_key),
                            field=field,
                            raw_value=raw_value,
                        )

                already_parsed = {
                    "contract_no",
                    "supply_center",
                    "revenue_month_rpd",
                    "revenue_month_cpd",
                    *_PREVIOUS_MANUAL_MONTH_FIELDS,
                }
                for column in config.base_columns:
                    field = column["id"]
                    if field in already_parsed:
                        continue
                    cell = field_cells[field]
                    raw_value = cell.value if cell is not None else None
                    if field in _PREVIOUS_MANUAL_AMOUNT_FIELDS:
                        value, valid = _parse_previous_manual_amount(cell)
                        values[field] = value
                        if not valid:
                            issues.add(
                                "INVALID_PREVIOUS_MANUAL_AMOUNT",
                                "上期人工金额非空且无法解析，已按空白处理",
                                workbook=workbook_path.name,
                                sheet=sheet.title,
                                row_number=row_number,
                                field=field,
                                raw_value=raw_value,
                            )
                    elif field in _PREVIOUS_MANUAL_FLAG_FIELDS:
                        value, valid = _parse_previous_manual_flag(cell)
                        values[field] = value
                        if not valid:
                            issues.add(
                                "INVALID_PREVIOUS_MANUAL_FLAG",
                                "上期人工标识只允许Y/N，已按空白处理",
                                workbook=workbook_path.name,
                                sheet=sheet.title,
                                row_number=row_number,
                                field=field,
                                raw_value=raw_value,
                            )
                    else:
                        values[field] = _parse_previous_value(
                            field,
                            raw_value,
                            workbook.epoch,
                        )
                key = business_key_identity(*display_key)
                if previous_row_kinds is None:
                    row_kind = (
                        CONTRACT_ONLY_NO_DEMAND
                        if bool(display_key[0])
                        and not display_key[1]
                        and normalize_text(values.get("revenue_segment"))
                        == "不要货"
                        else DEMAND_CENTER
                    )
                else:
                    row_kind = previous_row_kinds.get(key)
                    if row_kind is None:
                        issues.add(
                            "PREVIOUS_ROW_KIND_UNAVAILABLE",
                            "上期结果缺少该业务键的显式行状态，已排除该行",
                            workbook=workbook_path.name,
                            sheet="_tool_meta",
                            row_number=row_number,
                            business_key=" | ".join(display_key),
                            field="row_kind",
                        )
                        continue
                valid_key = bool(display_key[0]) and (
                    (
                        row_kind == CONTRACT_ONLY_NO_DEMAND
                        and not display_key[1]
                    )
                    or (row_kind == DEMAND_CENTER and bool(display_key[1]))
                )
                if not valid_key:
                    issues.add(
                        "PREVIOUS_INVALID_BUSINESS_KEY",
                        "上期基表业务键与显式行状态不一致，无法继承或比较",
                        workbook=workbook_path.name,
                        sheet=sheet.title,
                        row_number=row_number,
                        business_key=" | ".join(display_key),
                    )
                    continue
                if key in rows:
                    issues.add(
                        "PREVIOUS_DUPLICATE_BUSINESS_KEY",
                        "上期基表存在重复业务键，按原顺序保留第一条",
                        workbook=workbook_path.name,
                        sheet=sheet.title,
                        row_number=row_number,
                        business_key=_display_key(display_key),
                    )
                    continue
                rows[key] = BaseRow(values, row_kind=row_kind)
            return PreviousData(rows, usable=True)
        finally:
            workbook.close()

    def _read_role_sheet(
        self,
        workbook_path: Path,
        epoch: datetime,
        sheet,
        role: str,
        config: ToolConfig,
        issues: IssueLog,
        *,
        header_row: int,
    ) -> list[ParsedRow]:
        headers = [normalize_text(cell.value) for cell in sheet[header_row]]
        indexes: dict[str, int | None] = {}
        matched_by_index: dict[int, list[str]] = {}
        for field, field_spec in config.fields[role].items():
            match = resolve_name(
                field_spec["canonical"],
                field_spec.get("aliases", []),
                headers,
                config.workbook["contains_direction"],
            )
            indexes[field] = match.index
            if match.index is not None:
                matched_by_index.setdefault(match.index, []).append(field)
            elif match.mode == "ambiguous":
                candidate_names = [headers[index] for index in match.candidates]
                issues.add(
                    "AMBIGUOUS_FIELD",
                    f"字段 {field_spec['canonical']} 命中多个候选",
                    workbook=workbook_path.name,
                    sheet=sheet.title,
                    row_number=header_row,
                    field=field,
                    raw_value=" | ".join(candidate_names),
                )
            else:
                issues.add(
                    "MISSING_FIELD",
                    f"未找到字段: {field_spec['canonical']}",
                    workbook=workbook_path.name,
                    sheet=sheet.title,
                    row_number=header_row,
                    field=field,
                )

        for index, fields in matched_by_index.items():
            if len(fields) <= 1:
                continue
            for field in fields:
                indexes[field] = None
            issues.add(
                "FIELD_COLLISION",
                "一个源表头被多个内部字段占用，本次均视为不可用",
                workbook=workbook_path.name,
                sheet=sheet.title,
                row_number=header_row,
                field=" | ".join(fields),
                raw_value=headers[index],
            )

        seen_signatures: dict[tuple[tuple[str, str], ...], int] = {}
        result: list[ParsedRow] = []
        for row_number in range(header_row + 1, sheet.max_row + 1):
            cells = list(sheet[row_number])
            if _row_is_blank(cells):
                continue
            signature = tuple(
                normalize_signature_value(cell.value) for cell in cells
            )
            first_row = seen_signatures.get(signature)
            if first_row is not None:
                issues.add(
                    "DUPLICATE_ROW_IGNORED",
                    f"完全重复行已忽略，首次出现于第 {first_row} 行",
                    workbook=workbook_path.name,
                    sheet=sheet.title,
                    row_number=row_number,
                    raw_value=f"first_row={first_row}",
                )
                continue
            seen_signatures[signature] = row_number

            values: dict[str, Any] = {}
            raw_values: dict[str, Any] = {}
            invalid: set[str] = set()
            for field, field_spec in config.fields[role].items():
                index = indexes[field]
                cell = (
                    cells[index]
                    if index is not None and index < len(cells)
                    else None
                )
                raw = cell.value if cell is not None else None
                raw_values[field] = raw
                value, valid = _parse_source_cell(
                    field_spec["type"],
                    cell,
                    epoch,
                )
                values[field] = value
                if not valid:
                    invalid.add(field)
                    if not (role == "transit" and field == "transit_days"):
                        issues.add(
                            _error_code_for_type(field_spec["type"]),
                            _invalid_value_message(field_spec["type"]),
                            workbook=workbook_path.name,
                            sheet=sheet.title,
                            row_number=row_number,
                            field=field,
                            raw_value=raw,
                        )
            result.append(
                ParsedRow(
                    role=role,
                    workbook=workbook_path.name,
                    sheet=sheet.title,
                    row_number=row_number,
                    values=values,
                    raw_values=raw_values,
                    invalid_fields=frozenset(invalid),
                )
            )
        return result

    def _detect_output_header(
        self,
        sheet,
        config: ToolConfig,
        expected_names: list[str] | None = None,
    ) -> int | None:
        max_row = min(
            int(config.workbook["header_scan_rows"]),
            sheet.max_row,
        )
        expected = expected_names or [
            column["name"] for column in config.base_columns
        ]
        for row_number in range(1, max_row + 1):
            headers = [normalize_text(cell.value) for cell in sheet[row_number]]
            matched = sum(
                1
                for name in expected
                if resolve_name(
                    name,
                    [],
                    headers,
                    config.workbook["contains_direction"],
                ).index
                is not None
            )
            if matched >= min(5, len(expected)):
                return row_number
        return None


def _open_workbook(path: Path):
    if not path.exists():
        raise WorkbookReadError(f"工作簿不存在: {path}")
    try:
        return load_workbook(path, data_only=True, read_only=False)
    except Exception as exc:
        raise WorkbookReadError(f"工作簿无法读取: {path}: {exc}") from exc


def _missing_required_fields_summary(
    fingerprints,
    role: str,
    config: ToolConfig,
) -> str:
    parts: list[str] = []
    for fingerprint in fingerprints:
        missing = [
            (
                f"{field}"
                f"({config.fields[role][field]['canonical']})"
            )
            for field in fingerprint.missing_required_fields
        ]
        parts.append(
            f"{fingerprint.sheet_name}:"
            + (", ".join(missing) if missing else "无")
        )
    return " | ".join(parts) if parts else "无可扫描 Sheet"


def _row_is_blank(cells) -> bool:
    return all(
        is_business_blank(
            cell.value,
            data_type=getattr(cell, "data_type", None),
        )
        for cell in cells
    )


def _parse_source_cell(
    field_type: str, cell, epoch: datetime
) -> tuple[Any, bool]:
    raw = cell.value if cell is not None else None
    data_type = getattr(cell, "data_type", None) if cell is not None else None
    if is_business_blank(raw, data_type=data_type):
        return (ZERO_AMOUNT, True) if field_type == "amount" else (None, True)
    if data_type == "e":
        return (ZERO_AMOUNT, False) if field_type == "amount" else (None, False)
    if field_type == "text":
        return _text_from_cell(cell), True
    if field_type == "flag":
        value = normalize_text(raw).upper()
        return (value, True) if value in {"Y", "N"} else (None, False)
    if field_type == "amount":
        value = normalize_amount(raw)
        return (value, True) if value is not None else (ZERO_AMOUNT, False)
    if field_type == "date":
        value = _parse_date(raw, epoch)
        return value, value is not None
    if field_type == "nonnegative_integer":
        value = _parse_nonnegative_integer(
            raw,
            cell.number_format if cell is not None else None,
        )
        return value, value is not None
    return normalize_text(raw), True


def _text_from_cell(cell) -> str:
    raw = cell.value
    if isinstance(raw, bool):
        return "TRUE" if raw else "FALSE"
    if isinstance(raw, (int, float)) and not isinstance(raw, bool):
        number = Decimal(str(raw))
        text = (
            str(int(number))
            if number == number.to_integral_value()
            else format(number.normalize(), "f")
        )
        format_section = str(cell.number_format or "").split(";")[0]
        if number == number.to_integral_value() and re.fullmatch(
            r"0+", format_section
        ):
            return text.zfill(len(format_section))
        return text
    return normalize_text(raw)


def _parse_decimal(value: Any) -> Decimal | None:
    try:
        text = normalize_text(value).replace(",", "")
        negative = text.startswith("(") and text.endswith(")")
        if negative:
            text = text[1:-1]
        result = Decimal(text)
        if not result.is_finite():
            return None
        result = -result if negative else result
        as_float = float(result)
        if not math.isfinite(as_float) or (
            result != 0 and as_float == 0
        ):
            return None
        return result
    except (InvalidOperation, ValueError):
        return None


def _parse_nonnegative_integer(
    value: Any, number_format: str | None = None
) -> int | None:
    decimal = _parse_decimal(value)
    if decimal is None or decimal < 0:
        return None
    if decimal != decimal.to_integral_value():
        if not _displays_as_integer(number_format):
            return None
        decimal = decimal.quantize(Decimal("1"), rounding=ROUND_HALF_UP)
    return int(decimal)


def _displays_as_integer(number_format: str | None) -> bool:
    section = str(number_format or "").split(";", maxsplit=1)[0]
    section = re.sub(r"\[[^\]]*\]", "", section)
    section = re.sub(r'"[^"]*"', "", section)
    section = re.sub(r"_.", "", section)
    section = re.sub(r"\*.", "", section)
    section = section.replace("\\", "")
    section = "".join(section.split())
    return section == "#,##0"


def _parse_date(value: Any, epoch: datetime) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        try:
            converted = from_excel(value, epoch)
            if isinstance(converted, datetime):
                return converted.date()
            if isinstance(converted, date):
                return converted
            return None
        except (TypeError, ValueError, OverflowError):
            return None
    text = normalize_text(value)
    for pattern in (
        "%Y-%m-%d",
        "%Y/%m/%d",
        "%Y.%m.%d",
        "%Y年%m月%d日",
        "%Y%m%d",
        "%d-%b-%Y",
        "%d %b %Y",
    ):
        try:
            return datetime.strptime(text, pattern).date()
        except ValueError:
            continue
    return None


def _parse_previous_value(field: str, value: Any, epoch: datetime) -> Any:
    if field in {"legacy_amount", "monthly_new_order"}:
        amount = normalize_amount(value)
        return amount if amount is not None else ZERO_AMOUNT
    if is_business_blank(value):
        return None
    if field in {"contract_no", "supply_center"}:
        return normalize_text(value)
    if field in {
        "ata",
        "asd",
        "rpd",
        "latest_asd",
        "latest_rpd",
        "cpd",
        "arrival_date_rpd",
        "arrival_date_cpd",
    }:
        return _parse_date(value, epoch)
    if field in {"revenue_month_rpd", "revenue_month_cpd"}:
        if isinstance(value, (date, datetime)):
            return value.strftime("%Y-%m")
        return normalize_text(value)
    return normalize_text(value)


def _parse_previous_manual_amount(cell) -> tuple[Decimal | None, bool]:
    if cell is None:
        return None, True
    raw = cell.value
    data_type = getattr(cell, "data_type", None)
    if is_business_blank(raw, data_type=data_type):
        return None, True
    if data_type == "e":
        return None, False
    amount = normalize_amount(raw)
    return (amount, True) if amount is not None else (None, False)


def _manual_month_reference_context(
    primary_reference: Any,
    secondary_reference: Any,
) -> str:
    primary = normalize_text(primary_reference) or "空"
    secondary = normalize_text(secondary_reference) or "空"
    return (
        f"同口径自动收入年月={primary}；"
        f"另一口径自动收入年月={secondary}"
    )


def _parse_previous_manual_flag(cell) -> tuple[str | None, bool]:
    if cell is None:
        return None, True
    data_type = getattr(cell, "data_type", None)
    if is_business_blank(cell.value, data_type=data_type):
        return None, True
    if data_type == "e":
        return None, False
    value = normalize_text(cell.value).upper()
    return (value, True) if value in {"Y", "N"} else (None, False)


def _error_code_for_type(field_type: str) -> str:
    return {
        "amount": "INVALID_AMOUNT",
        "date": "INVALID_DATE",
        "flag": "INVALID_ENUM_VALUE",
        "nonnegative_integer": "INVALID_TRANSIT_DAYS",
    }.get(field_type, "INVALID_VALUE")


def _invalid_value_message(field_type: str) -> str:
    if field_type == "amount":
        return "非空金额无法解析，已记录异常并按数值0继续"
    return f"字段值无法按 {field_type} 解析，已排除该值"


def _display_key(key: tuple[str, str]) -> str:
    return f"{key[0]} | {key[1]}"


def _read_previous_metadata(
    workbook,
    config: ToolConfig,
    workbook_path: Path,
    issues: IssueLog,
) -> tuple[
    str,
    dict[str, str],
    dict[tuple[str, str], str] | None,
]:
    default_sheet = config.output["sheets"]["base"]
    default_names = config.base_names_by_id
    if "_tool_meta" not in workbook.sheetnames:
        return default_sheet, default_names, None
    sheet = workbook["_tool_meta"]
    try:
        if (
            normalize_text(sheet["A1"].value) != "schema_version"
            or normalize_text(sheet["A2"].value) != "base_sheet"
            or normalize_text(sheet["A4"].value) != "field_id"
        ):
            raise ValueError("metadata header mismatch")
        schema_version = normalize_text(sheet["B1"].value)
        if schema_version not in {"2", "3"}:
            raise ValueError("unsupported metadata schema")
        base_sheet = normalize_text(sheet["B2"].value)
        names: dict[str, str] = {}
        for row_number in range(5, sheet.max_row + 1):
            field = normalize_text(sheet.cell(row_number, 1).value)
            name = normalize_text(sheet.cell(row_number, 2).value)
            if not field:
                break
            if field not in config.base_names_by_id:
                raise ValueError("unknown field metadata")
            if not name:
                raise ValueError("field metadata name missing")
            names[field] = name
        if not base_sheet or not names:
            raise ValueError("metadata content is incomplete")
        if schema_version == "2":
            return base_sheet, names, None

        row_kind_header = None
        for row_number in range(5, sheet.max_row + 1):
            if (
                normalize_text(sheet.cell(row_number, 1).value)
                == "row_kind_contract_no"
                and normalize_text(sheet.cell(row_number, 2).value)
                == "row_kind_supply_center"
                and normalize_text(sheet.cell(row_number, 3).value)
                == "row_kind"
            ):
                row_kind_header = row_number
                break
        if row_kind_header is None:
            raise ValueError("row kind metadata missing")
        row_kinds: dict[tuple[str, str], str] = {}
        for row_number in range(row_kind_header + 1, sheet.max_row + 1):
            contract_no = normalize_text(sheet.cell(row_number, 1).value)
            supply_center = normalize_text(sheet.cell(row_number, 2).value)
            row_kind = normalize_text(sheet.cell(row_number, 3).value)
            if not contract_no and not supply_center and not row_kind:
                continue
            if row_kind not in {DEMAND_CENTER, CONTRACT_ONLY_NO_DEMAND}:
                raise ValueError("invalid row kind")
            if not contract_no or (
                row_kind == DEMAND_CENTER and not supply_center
            ) or (
                row_kind == CONTRACT_ONLY_NO_DEMAND and supply_center
            ):
                raise ValueError("row kind business key mismatch")
            key = business_key_identity(contract_no, supply_center)
            if key in row_kinds:
                raise ValueError("duplicate row kind business key")
            row_kinds[key] = row_kind
        return base_sheet, names, row_kinds
    except Exception:
        issues.add(
            "PREVIOUS_METADATA_INVALID",
            "上期结果的内部字段元数据无效，回退到当前显示名匹配",
            workbook=workbook_path.name,
            sheet="_tool_meta",
        )
        return default_sheet, default_names, None
