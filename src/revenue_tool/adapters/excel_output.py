from __future__ import annotations

from datetime import datetime, timezone
from decimal import Decimal
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from revenue_tool.domain.models import (
    ComparisonLine,
    RevenueLine,
    RevenueSummaryLine,
)


class ExcelOutputAdapter:
    def write(
        self,
        path: str | Path,
        summary: list[RevenueSummaryLine],
        details: list[RevenueLine],
        comparison: list[ComparisonLine],
        source_file: str,
        previous_file: str | None,
    ) -> None:
        workbook = Workbook()
        workbook.remove(workbook.active)
        self._write_summary(workbook, summary)
        self._write_details(workbook, details)
        self._write_comparison(workbook, comparison)
        self._write_run_info(workbook, source_file, previous_file, len(details))

        output_path = Path(path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)
        workbook.close()

    def _write_summary(self, workbook, rows: list[RevenueSummaryLine]) -> None:
        headers = [
            "Revenue Month",
            "Contract Number",
            "Shipping Point",
            "Trade Type",
            "Plan Quantity",
            "Revenue Amount",
            "Shipment Count",
        ]
        sheet = workbook.create_sheet("Revenue Summary")
        _append_table(
            sheet,
            headers,
            (
                (
                    row.revenue_month,
                    row.contract_number,
                    row.shipping_point,
                    row.trade_type,
                    _excel_number(row.plan_quantity),
                    _excel_number(row.revenue_amount),
                    row.shipment_count,
                )
                for row in rows
            ),
        )

    def _write_details(self, workbook, rows: list[RevenueLine]) -> None:
        headers = [
            "Business Key",
            "PO Number",
            "Contract Number",
            "Shipping Point",
            "Shipment ID",
            "Trade Type",
            "PRD",
            "Original PO Quantity",
            "Plan Date",
            "Plan Quantity",
            "Transit Days",
            "Arrival Date",
            "Revenue Month",
            "Revenue Amount",
        ]
        sheet = workbook.create_sheet("Revenue Detail")
        _append_table(
            sheet,
            headers,
            (
                (
                    row.business_key,
                    row.po_number,
                    row.contract_number,
                    row.shipping_point,
                    row.shipment_id,
                    row.trade_type,
                    row.prd,
                    _excel_number(row.original_po_quantity),
                    row.plan_date,
                    _excel_number(row.plan_quantity),
                    row.transit_days,
                    row.arrival_date,
                    row.revenue_month,
                    _excel_number(row.revenue_amount),
                )
                for row in rows
            ),
        )
        for column in ("G", "I", "L"):
            for cell in sheet[column][1:]:
                cell.number_format = "yyyy-mm-dd"

    def _write_comparison(self, workbook, rows: list[ComparisonLine]) -> None:
        headers = [
            "Business Key",
            "PO Number",
            "Contract Number",
            "Shipping Point",
            "Shipment ID",
            "Previous Revenue Month",
            "Current Revenue Month",
            "Delay Months",
            "Status",
        ]
        sheet = workbook.create_sheet("Comparison")
        _append_table(
            sheet,
            headers,
            (
                (
                    row.business_key,
                    row.po_number,
                    row.contract_number,
                    row.shipping_point,
                    row.shipment_id,
                    row.previous_revenue_month,
                    row.current_revenue_month,
                    row.delay_months,
                    "Delayed" if row.delayed else "Not Delayed",
                )
                for row in rows
            ),
        )
        for cell in sheet["I"][1:]:
            if cell.value == "Delayed":
                cell.fill = PatternFill("solid", fgColor="FFC7CE")
                cell.font = Font(color="9C0006", bold=True)

    def _write_run_info(
        self,
        workbook,
        source_file: str,
        previous_file: str | None,
        detail_count: int,
    ) -> None:
        sheet = workbook.create_sheet("Run Info")
        rows = [
            ("Generated At (UTC)", datetime.now(timezone.utc).isoformat(timespec="seconds")),
            ("Source File", source_file),
            ("Previous Result", previous_file or ""),
            ("Revenue Detail Count", detail_count),
            ("Tool Version", "0.1.0"),
        ]
        for row in rows:
            sheet.append(row)
        sheet.column_dimensions["A"].width = 24
        sheet.column_dimensions["B"].width = 60


def _append_table(sheet, headers: list[str], rows: Iterable[tuple]) -> None:
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for index, header in enumerate(headers, start=1):
        values = [len(str(header))]
        values.extend(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in list(sheet.columns)[index - 1][1:]
        )
        sheet.column_dimensions[get_column_letter(index)].width = min(max(values) + 2, 42)


def _excel_number(value: Decimal | None):
    if value is None:
        return None
    if value == value.to_integral_value():
        return int(value)
    return float(value)

